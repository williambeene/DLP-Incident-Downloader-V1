"""
DLP 16.x Incident Downloader
A Python UI application for downloading and exporting DLP incident details
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext, simpledialog
import requests
from requests.adapters import HTTPAdapter
import json
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
import urllib3
import ssl
from datetime import datetime
import os
import csv
import queue
import io
import logging
import tempfile
import shutil

# Optional text extraction libraries
try:
    import PyPDF2
    HAS_PYPDF2 = True
except ImportError:
    HAS_PYPDF2 = False

try:
    import docx
    HAS_DOCX = True
except ImportError:
    HAS_DOCX = False

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Disable SSL warnings for self-signed certificates
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


# Magic byte signatures for file type detection
MAGIC_BYTES = {
    # Documents
    b'%PDF': '.pdf',
    b'PK\x03\x04': '.zip',  # Also used by docx, xlsx, pptx - will check further
    b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1': '.doc',  # OLE compound document (doc, xls, ppt)
    b'{\\rtf': '.rtf',

    # Images
    b'\xff\xd8\xff': '.jpg',
    b'\x89PNG\r\n\x1a\n': '.png',
    b'GIF87a': '.gif',
    b'GIF89a': '.gif',
    b'BM': '.bmp',
    b'II*\x00': '.tiff',  # Little-endian TIFF
    b'MM\x00*': '.tiff',  # Big-endian TIFF
    b'RIFF': '.webp',  # Could also be .wav, .avi - check further bytes

    # Archives
    b'Rar!\x1a\x07': '.rar',
    b'7z\xbc\xaf\x27\x1c': '.7z',
    b'\x1f\x8b': '.gz',
    b'BZh': '.bz2',

    # Executables
    b'MZ': '.exe',
    b'\x7fELF': '.elf',

    # Audio/Video
    b'ID3': '.mp3',
    b'\xff\xfb': '.mp3',
    b'\xff\xfa': '.mp3',
    b'OggS': '.ogg',
    b'ftyp': '.mp4',  # Actually starts at offset 4
    b'\x00\x00\x00\x1cftyp': '.mp4',
    b'\x00\x00\x00\x20ftyp': '.mp4',

    # Email
    b'From:': '.eml',
    b'Return-Path:': '.eml',
    b'Received:': '.eml',
    b'MIME-Version:': '.eml',

    # XML/HTML
    b'<?xml': '.xml',
    b'<!DOCTYPE html': '.html',
    b'<html': '.html',
    b'<HTML': '.html',
}


def detect_file_type_by_magic(data):
    """Detect file type by examining magic bytes at the start of the file"""
    if not data or len(data) < 8:
        return None

    # Check for ZIP-based formats (docx, xlsx, pptx, etc.)
    if data[:4] == b'PK\x03\x04':
        # Need to check internal structure to determine Office format
        try:
            import zipfile
            from io import BytesIO
            with zipfile.ZipFile(BytesIO(data), 'r') as zf:
                names = zf.namelist()
                if any('word/' in n for n in names):
                    return '.docx'
                elif any('xl/' in n for n in names):
                    return '.xlsx'
                elif any('ppt/' in n for n in names):
                    return '.pptx'
                elif 'META-INF/MANIFEST.MF' in names:
                    return '.jar'
                else:
                    return '.zip'
        except:
            return '.zip'

    # Check for OLE compound documents (older Office formats)
    if data[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
        # This could be .doc, .xls, .ppt, .msg
        # Check for specific markers
        if b'Word.Document' in data[:4096] or b'MSWordDoc' in data[:4096]:
            return '.doc'
        elif b'Microsoft Excel' in data[:4096] or b'Workbook' in data[:4096]:
            return '.xls'
        elif b'PowerPoint' in data[:4096]:
            return '.ppt'
        elif b'\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00' in data[:32]:
            # Could be MSG file
            return '.msg'
        return '.doc'  # Default to .doc for OLE files

    # Check MP4/MOV (ftyp at offset 4)
    if len(data) >= 12 and data[4:8] == b'ftyp':
        ftyp_brand = data[8:12]
        if ftyp_brand in (b'isom', b'mp41', b'mp42', b'avc1', b'M4V ', b'M4A '):
            return '.mp4'
        elif ftyp_brand == b'qt  ':
            return '.mov'
        return '.mp4'

    # Check RIFF formats (wav, avi, webp)
    if data[:4] == b'RIFF' and len(data) >= 12:
        riff_type = data[8:12]
        if riff_type == b'WAVE':
            return '.wav'
        elif riff_type == b'AVI ':
            return '.avi'
        elif riff_type == b'WEBP':
            return '.webp'

    # Standard magic byte checks
    for magic, ext in MAGIC_BYTES.items():
        if data[:len(magic)] == magic:
            return ext

    # Check if it's plain text or email
    try:
        # Try to decode as text
        sample = data[:4096]
        text_sample = sample.decode('utf-8', errors='ignore').strip()
        text_lower = text_sample.lower()

        # Check for email indicators (EML format)
        email_headers = ['from:', 'to:', 'subject:', 'date:', 'received:', 'return-path:',
                        'mime-version:', 'content-type:', 'message-id:', 'x-mailer:',
                        'delivered-to:', 'reply-to:', 'cc:', 'bcc:']
        header_count = sum(1 for h in email_headers if h in text_lower)

        # If we find multiple email headers, it's likely an EML file
        if header_count >= 2:
            return '.eml'

        # Check for common text patterns
        if b',' in sample and b'\n' in sample and header_count == 0:
            # Could be CSV - check for consistent comma-separated structure
            lines = text_sample.split('\n')[:5]
            if len(lines) > 1:
                comma_counts = [line.count(',') for line in lines if line.strip()]
                if len(comma_counts) > 1 and len(set(comma_counts)) <= 2:  # Consistent comma count
                    return '.csv'
        elif b'\t' in sample and b'\n' in sample:
            return '.tsv'

        # Default to .txt for plain text
        return '.txt'
    except:
        pass

    return None


class TLSAdapter(HTTPAdapter):
    """Custom adapter to handle older TLS versions and cipher suites"""
    def init_poolmanager(self, *args, **kwargs):
        # Create a custom SSL context that's more permissive
        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE
        # Allow older TLS versions
        ctx.set_ciphers('DEFAULT:@SECLEVEL=1')
        ctx.options &= ~ssl.OP_NO_SSLv3
        kwargs['ssl_context'] = ctx
        return super().init_poolmanager(*args, **kwargs)


class DLPIncidentDownloader:
    """Main application class for DLP Incident Downloader"""

    # Environment configurations
    ENVIRONMENTS = {
        "DAR": {"name": "Data At Rest (DAR)", "description": "Network Discover/Protect"},
        "WEB": {"name": "Web (Network Prevent)", "description": "Network Prevent for Web"},
        "MAIL": {"name": "Mail (Network Prevent)", "description": "Network Prevent for Email"},
        "ENDPOINT": {"name": "Endpoint", "description": "Endpoint Prevent/Discover"}
    }

    # Config file path
    CONFIG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "dlp_config.json")

    def __init__(self, root):
        self.root = root
        self.root.title("DLP 16.x Incident Downloader")
        self.root.geometry("1000x750")
        self.root.minsize(900, 650)

        # Setup logging
        self._setup_logging()

        # Session and connection state
        self.session = None
        self.base_url = None
        self.is_connected = False
        self.current_incidents = []

        # File-based incident cache (instead of in-memory dict)
        self._cache_dir = None
        self._fetched_incident_ids = set()

        # Saved environments
        self.saved_environments = {}

        # Fetch state for stop/resume
        self._stop_fetch = False
        self._fetch_in_progress = False
        self._resume_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "dlp_resume_state.json")

        # Create the UI
        self.create_ui()

        # Load saved config
        self.load_config()

        # Cleanup on close
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

    def _setup_logging(self):
        """Setup file and console logging"""
        app_dir = os.path.dirname(os.path.abspath(__file__))
        log_dir = os.path.join(app_dir, "logs")
        os.makedirs(log_dir, exist_ok=True)

        log_file = os.path.join(log_dir, f"dlp_downloader_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")

        # Create logger
        self.logger = logging.getLogger("DLPDownloader")
        self.logger.setLevel(logging.DEBUG)

        # File handler - detailed logging
        file_handler = logging.FileHandler(log_file, encoding='utf-8')
        file_handler.setLevel(logging.DEBUG)
        file_format = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        file_handler.setFormatter(file_format)
        self.logger.addHandler(file_handler)

        self.logger.info(f"DLP Incident Downloader started - Log file: {log_file}")

    def _init_cache_dir(self):
        """Initialize temporary directory for caching incident data"""
        if self._cache_dir and os.path.exists(self._cache_dir):
            return  # Already initialized

        # Create temp directory for this session
        self._cache_dir = tempfile.mkdtemp(prefix="dlp_cache_")
        self._fetched_incident_ids = set()
        self.logger.info(f"Initialized cache directory: {self._cache_dir}")

    def _cleanup_cache(self):
        """Clean up temporary cache directory"""
        if self._cache_dir and os.path.exists(self._cache_dir):
            try:
                shutil.rmtree(self._cache_dir)
                self.logger.info(f"Cleaned up cache directory: {self._cache_dir}")
            except Exception as e:
                self.logger.error(f"Failed to cleanup cache: {e}")
            self._cache_dir = None
            self._fetched_incident_ids = set()

    def _save_incident_to_cache(self, incident_id, details):
        """Save incident details to cache file"""
        if not self._cache_dir:
            self._init_cache_dir()

        cache_file = os.path.join(self._cache_dir, f"incident_{incident_id}.json")
        try:
            with open(cache_file, 'w', encoding='utf-8') as f:
                json.dump(details, f, default=str)
            self._fetched_incident_ids.add(str(incident_id))
            return True
        except Exception as e:
            self.logger.error(f"Failed to cache incident {incident_id}: {e}")
            return False

    def _load_incident_from_cache(self, incident_id):
        """Load incident details from cache file"""
        if not self._cache_dir:
            return None

        cache_file = os.path.join(self._cache_dir, f"incident_{incident_id}.json")
        try:
            if os.path.exists(cache_file):
                with open(cache_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except Exception as e:
            self.logger.error(f"Failed to load incident {incident_id} from cache: {e}")
        return None

    def _get_cached_incident_ids(self):
        """Get list of cached incident IDs"""
        return self._fetched_incident_ids.copy()

    def _iter_cached_incidents(self):
        """Generator to iterate through cached incidents without loading all into memory"""
        if not self._cache_dir:
            return

        for incident_id in self._fetched_incident_ids:
            details = self._load_incident_from_cache(incident_id)
            if details:
                yield incident_id, details

    def _on_close(self):
        """Handle application close"""
        if self._fetch_in_progress:
            if not messagebox.askyesno("Confirm Exit",
                                       "A download is in progress. Are you sure you want to exit?\n\n"
                                       "Progress will be saved and can be resumed later."):
                return
            self._stop_fetch = True

        # Cleanup cache if no resume needed
        if not self._fetch_in_progress:
            self._cleanup_cache()

        self.logger.info("Application closed")
        self.root.destroy()

    def create_ui(self):
        """Create the main user interface"""
        # Create main container with padding
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky="nsew")

        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(3, weight=1)

        # === Connection Section ===
        self.create_connection_section(main_frame)

        # === Report Section ===
        self.create_report_section(main_frame)

        # === Progress Section ===
        self.create_progress_section(main_frame)

        # === Results Section ===
        self.create_results_section(main_frame)

        # === Export Section ===
        self.create_export_section(main_frame)

    def create_connection_section(self, parent):
        """Create the connection/login section"""
        conn_frame = ttk.LabelFrame(parent, text="Connection Settings", padding="10")
        conn_frame.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        conn_frame.columnconfigure(1, weight=1)

        # Saved Environments dropdown
        ttk.Label(conn_frame, text="Saved Environment:").grid(row=0, column=0, sticky="w", padx=(0, 10))
        self.saved_env_var = tk.StringVar(value="")
        self.saved_env_combo = ttk.Combobox(conn_frame, textvariable=self.saved_env_var,
                                             values=[], state="readonly", width=30)
        self.saved_env_combo.grid(row=0, column=1, sticky="w")
        self.saved_env_combo.bind("<<ComboboxSelected>>", self.on_saved_environment_select)

        # Save/Delete environment buttons
        env_btn_frame = ttk.Frame(conn_frame)
        env_btn_frame.grid(row=0, column=2, sticky="e", padx=(10, 0))
        ttk.Button(env_btn_frame, text="Save", width=8, command=self.save_environment).pack(side="left", padx=(0, 5))
        ttk.Button(env_btn_frame, text="Delete", width=8, command=self.delete_environment).pack(side="left")

        # Environment Type selection
        ttk.Label(conn_frame, text="Environment Type:").grid(row=1, column=0, sticky="w", padx=(0, 10), pady=(10, 0))
        self.env_var = tk.StringVar(value="ENDPOINT")
        env_combo = ttk.Combobox(conn_frame, textvariable=self.env_var,
                                  values=list(self.ENVIRONMENTS.keys()),
                                  state="readonly", width=15)
        env_combo.grid(row=1, column=1, sticky="w", pady=(10, 0))
        env_combo.bind("<<ComboboxSelected>>", self.on_environment_change)

        # Environment description label
        self.env_desc_label = ttk.Label(conn_frame, text=self.ENVIRONMENTS["ENDPOINT"]["description"],
                                        foreground="gray")
        self.env_desc_label.grid(row=1, column=2, sticky="w", padx=(10, 0), pady=(10, 0))

        # Enforce Server URL
        ttk.Label(conn_frame, text="Enforce Server:").grid(row=2, column=0, sticky="w", padx=(0, 10), pady=(10, 0))
        self.server_entry = ttk.Entry(conn_frame, width=50)
        self.server_entry.grid(row=2, column=1, columnspan=2, sticky="ew", pady=(10, 0))
        self.server_entry.insert(0, "https://")

        # Username
        ttk.Label(conn_frame, text="Username:").grid(row=3, column=0, sticky="w", padx=(0, 10), pady=(10, 0))
        self.username_entry = ttk.Entry(conn_frame, width=30)
        self.username_entry.grid(row=3, column=1, sticky="w", pady=(10, 0))
        self.username_entry.insert(0, "Administrator")

        # Password
        ttk.Label(conn_frame, text="Password:").grid(row=4, column=0, sticky="w", padx=(0, 10), pady=(10, 0))
        self.password_entry = ttk.Entry(conn_frame, width=30, show="*")
        self.password_entry.grid(row=4, column=1, sticky="w", pady=(10, 0))

        # Connect button
        self.connect_btn = ttk.Button(conn_frame, text="Connect", command=self.connect)
        self.connect_btn.grid(row=4, column=2, sticky="e", pady=(10, 0), padx=(10, 0))

        # Connection status
        self.conn_status_label = ttk.Label(conn_frame, text="Not Connected", foreground="red")
        self.conn_status_label.grid(row=5, column=0, columnspan=3, sticky="w", pady=(10, 0))

    def create_report_section(self, parent):
        """Create the report ID and incident ID input section"""
        report_frame = ttk.LabelFrame(parent, text="Incident Source", padding="10")
        report_frame.grid(row=1, column=0, sticky="ew", pady=(0, 10))
        report_frame.columnconfigure(1, weight=1)

        # Source type selection
        self.source_type_var = tk.StringVar(value="report")
        source_frame = ttk.Frame(report_frame)
        source_frame.grid(row=0, column=0, columnspan=7, sticky="w", pady=(0, 10))

        ttk.Radiobutton(source_frame, text="Saved Report", variable=self.source_type_var,
                       value="report", command=self._on_source_type_change).pack(side="left", padx=(0, 20))
        ttk.Radiobutton(source_frame, text="Incident IDs", variable=self.source_type_var,
                       value="incidents", command=self._on_source_type_change).pack(side="left")

        # Report ID row
        self.report_row_frame = ttk.Frame(report_frame)
        self.report_row_frame.grid(row=1, column=0, columnspan=7, sticky="ew")

        ttk.Label(self.report_row_frame, text="Report ID:").pack(side="left", padx=(0, 10))
        self.report_id_entry = ttk.Entry(self.report_row_frame, width=20)
        self.report_id_entry.pack(side="left")

        # Fetch Report button
        self.fetch_report_btn = ttk.Button(self.report_row_frame, text="Fetch Report",
                                           command=self.fetch_report, state="disabled")
        self.fetch_report_btn.pack(side="left", padx=(10, 0))

        # Incident IDs row (initially hidden)
        self.incident_ids_row_frame = ttk.Frame(report_frame)

        ttk.Label(self.incident_ids_row_frame, text="Incident IDs:").pack(side="left", padx=(0, 10))
        self.incident_ids_entry = ttk.Entry(self.incident_ids_row_frame, width=50)
        self.incident_ids_entry.pack(side="left")
        ttk.Label(self.incident_ids_row_frame, text="(comma-separated or ranges: 1,2,3 or 100-200)",
                 foreground="gray").pack(side="left", padx=(10, 0))

        # Load Incidents button
        self.load_incidents_btn = ttk.Button(self.incident_ids_row_frame, text="Load Incidents",
                                             command=self.load_incident_ids, state="disabled")
        self.load_incidents_btn.pack(side="left", padx=(10, 0))

        # Action buttons row
        btn_frame = ttk.Frame(report_frame)
        btn_frame.grid(row=2, column=0, columnspan=7, sticky="w", pady=(10, 0))

        # Fetch All Details button
        self.fetch_details_btn = ttk.Button(btn_frame, text="Fetch All Incident Details",
                                            command=self.fetch_all_incident_details, state="disabled")
        self.fetch_details_btn.pack(side="left", padx=(0, 10))

        # Test Download button (fetch limited number)
        self.test_download_btn = ttk.Button(btn_frame, text="Test Download",
                                            command=self.test_download, state="disabled")
        self.test_download_btn.pack(side="left", padx=(0, 10))

        # Stop button
        self.stop_btn = ttk.Button(btn_frame, text="Stop", command=self.stop_fetch, state="disabled")
        self.stop_btn.pack(side="left", padx=(0, 10))

        # Resume button
        self.resume_btn = ttk.Button(btn_frame, text="Resume", command=self.resume_fetch, state="disabled")
        self.resume_btn.pack(side="left")

    def create_progress_section(self, parent):
        """Create the progress indicator section"""
        progress_frame = ttk.Frame(parent)
        progress_frame.grid(row=2, column=0, sticky="ew", pady=(0, 10))
        progress_frame.columnconfigure(0, weight=1)

        # Progress bar
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(progress_frame, variable=self.progress_var,
                                            maximum=100, mode='determinate')
        self.progress_bar.grid(row=0, column=0, sticky="ew")

        # Progress label
        self.progress_label = ttk.Label(progress_frame, text="Ready")
        self.progress_label.grid(row=1, column=0, sticky="w", pady=(5, 0))

    def create_results_section(self, parent):
        """Create the results display section"""
        results_frame = ttk.LabelFrame(parent, text="Results / Log", padding="10")
        results_frame.grid(row=3, column=0, sticky="nsew", pady=(0, 10))
        results_frame.columnconfigure(0, weight=1)
        results_frame.rowconfigure(0, weight=1)

        # Results text area with scrollbar
        self.results_text = scrolledtext.ScrolledText(results_frame, wrap=tk.WORD,
                                                       height=15, font=("Consolas", 9))
        self.results_text.grid(row=0, column=0, sticky="nsew")

    def create_export_section(self, parent):
        """Create the export section"""
        export_frame = ttk.LabelFrame(parent, text="Export", padding="10")
        export_frame.grid(row=4, column=0, sticky="ew")
        export_frame.columnconfigure(1, weight=1)

        # Export format selection
        ttk.Label(export_frame, text="Format:").grid(row=0, column=0, sticky="w", padx=(0, 10))
        self.export_format_var = tk.StringVar(value="JSON + Attachments")
        format_combo = ttk.Combobox(export_frame, textvariable=self.export_format_var,
                                    values=["JSON + Attachments", "JSON (AI-Friendly)", "CSV (Summary)", "Individual JSON Files"],
                                    state="readonly", width=20)
        format_combo.grid(row=0, column=1, sticky="w")

        # Incident count label
        self.incident_count_label = ttk.Label(export_frame, text="No incidents loaded")
        self.incident_count_label.grid(row=0, column=2, padx=(20, 0))

        # Output path
        ttk.Label(export_frame, text="Output Path:").grid(row=1, column=0, sticky="w", padx=(0, 10), pady=(10, 0))
        self.output_path_entry = ttk.Entry(export_frame, width=60)
        self.output_path_entry.grid(row=1, column=1, sticky="ew", pady=(10, 0))
        # Default to desktop
        default_path = os.path.join(os.path.expanduser("~"), "Desktop", "dlp_incidents_export.json")
        self.output_path_entry.insert(0, default_path)

        # Browse button
        self.browse_btn = ttk.Button(export_frame, text="Browse...", command=self.browse_export_path)
        self.browse_btn.grid(row=1, column=2, padx=(10, 0), pady=(10, 0))

        # Export button
        self.export_btn = ttk.Button(export_frame, text="Export Incidents",
                                     command=self.export_incidents, state="disabled")
        self.export_btn.grid(row=1, column=3, padx=(10, 0), pady=(10, 0))

    def browse_export_path(self):
        """Open file browser for export path"""
        export_format = self.export_format_var.get()

        try:
            if export_format == "Individual JSON Files":
                # For individual files, select a folder
                path = filedialog.askdirectory(parent=self.root, title="Select Export Folder")
                if path:
                    self.output_path_entry.delete(0, tk.END)
                    self.output_path_entry.insert(0, path)
            else:
                # For single file exports
                ext = ".csv" if "CSV" in export_format else ".json"
                ftypes = [("CSV files", "*.csv")] if "CSV" in export_format else [("JSON files", "*.json")]
                ftypes.append(("All files", "*.*"))

                path = filedialog.asksaveasfilename(
                    parent=self.root,
                    defaultextension=ext,
                    filetypes=ftypes
                )
                if path:
                    self.output_path_entry.delete(0, tk.END)
                    self.output_path_entry.insert(0, path)
        except Exception as e:
            self.log(f"Browse dialog error: {str(e)}", "ERROR")

    def on_environment_change(self, event=None):
        """Handle environment selection change"""
        env = self.env_var.get()
        if env in self.ENVIRONMENTS:
            self.env_desc_label.config(text=self.ENVIRONMENTS[env]["description"])

    def on_saved_environment_select(self, event=None):
        """Handle saved environment selection"""
        env_name = self.saved_env_var.get()
        if env_name and env_name in self.saved_environments:
            env_data = self.saved_environments[env_name]
            # Populate fields
            self.server_entry.delete(0, tk.END)
            self.server_entry.insert(0, env_data.get("server", "https://"))

            self.username_entry.delete(0, tk.END)
            self.username_entry.insert(0, env_data.get("username", "Administrator"))

            # Password not loaded from config for security - user must enter each time
            self.password_entry.delete(0, tk.END)
            # self.password_entry.insert(0, env_data.get("password", ""))

            env_type = env_data.get("env_type", "ENDPOINT")
            if env_type in self.ENVIRONMENTS:
                self.env_var.set(env_type)
                self.on_environment_change()

            self.log(f"Loaded environment: {env_name}")

    def save_environment(self):
        """Save current environment settings"""
        server = self.server_entry.get().strip()
        username = self.username_entry.get().strip()
        password = self.password_entry.get()
        env_type = self.env_var.get()

        if not server or server == "https://":
            messagebox.showerror("Error", "Please enter a server URL before saving")
            return

        # Ask for environment name
        current_name = self.saved_env_var.get()
        env_name = tk.simpledialog.askstring(
            "Save Environment",
            "Enter a name for this environment:",
            initialvalue=current_name if current_name else "",
            parent=self.root
        )

        if not env_name:
            return

        # Save environment data (password NOT saved for security)
        self.saved_environments[env_name] = {
            "server": server,
            "username": username,
            # "password": password,  # SECURITY: Password not stored in config
            "env_type": env_type
        }

        # Update dropdown
        self.update_saved_env_dropdown()
        self.saved_env_var.set(env_name)

        # Save to config file
        self.save_config()
        self.log(f"Saved environment: {env_name}", "SUCCESS")

    def delete_environment(self):
        """Delete selected saved environment"""
        env_name = self.saved_env_var.get()
        if not env_name:
            messagebox.showinfo("Info", "No environment selected to delete")
            return

        if messagebox.askyesno("Confirm Delete", f"Delete environment '{env_name}'?"):
            if env_name in self.saved_environments:
                del self.saved_environments[env_name]
                self.update_saved_env_dropdown()
                self.saved_env_var.set("")
                self.save_config()
                self.log(f"Deleted environment: {env_name}")

    def update_saved_env_dropdown(self):
        """Update the saved environments dropdown"""
        env_names = sorted(self.saved_environments.keys())
        self.saved_env_combo['values'] = env_names

    def save_config(self):
        """Save configuration to file"""
        try:
            config = {
                "saved_environments": self.saved_environments
            }
            with open(self.CONFIG_FILE, 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=2)
        except Exception as e:
            self.log(f"Error saving config: {str(e)}", "ERROR")

    def load_config(self):
        """Load configuration from file"""
        try:
            if os.path.exists(self.CONFIG_FILE):
                with open(self.CONFIG_FILE, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                self.saved_environments = config.get("saved_environments", {})
                self.update_saved_env_dropdown()
                self.log(f"Loaded {len(self.saved_environments)} saved environments")

            # Check for resume state
            self._update_resume_button()
        except Exception as e:
            self.log(f"Error loading config: {str(e)}", "ERROR")

    def log(self, message, level="INFO"):
        """Add a message to the results log and file logger"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        formatted_msg = f"[{timestamp}] [{level}] {message}\n"
        self.results_text.insert(tk.END, formatted_msg)
        self.results_text.see(tk.END)
        self.root.update_idletasks()

        # Also log to file
        if hasattr(self, 'logger'):
            log_level = getattr(logging, level.upper(), logging.INFO)
            self.logger.log(log_level, message)

    def update_progress(self, value, message=""):
        """Update the progress bar and label"""
        self.progress_var.set(value)
        if message:
            self.progress_label.config(text=message)
        self.root.update_idletasks()

    def get_base_url(self):
        """Get the base URL for API calls"""
        server = self.server_entry.get().strip()
        if not server.startswith("https://"):
            server = "https://" + server.replace("http://", "")
        # Remove trailing slash
        server = server.rstrip("/")
        return f"{server}/ProtectManager/webservices/v2"

    def connect(self):
        """Connect to the Enforce server"""
        if self.is_connected:
            self.disconnect()
            return

        server = self.server_entry.get().strip()
        username = self.username_entry.get().strip()
        password = self.password_entry.get()

        if not server or not username or not password:
            messagebox.showerror("Error", "Please fill in all connection fields")
            return

        self.log(f"Connecting to {server}...")
        self.update_progress(0, "Connecting...")

        # Run connection in thread to avoid UI freeze
        thread = threading.Thread(target=self._connect_thread, args=(server, username, password))
        thread.daemon = True
        thread.start()

    def _connect_thread(self, server, username, password):
        """Thread function for connection"""
        try:
            self.base_url = self.get_base_url()

            self.root.after(0, lambda: self.log(f"Attempting to connect to: {self.base_url}"))

            # Create session with basic auth
            self.session = requests.Session()
            self.session.auth = (username, password)
            self.session.verify = False  # Skip SSL verification for self-signed certs

            # Mount custom TLS adapter for older servers
            self.session.mount('https://', TLSAdapter())

            self.session.headers.update({
                "Content-Type": "application/json",
                "Accept": "application/json"
            })

            # Test connection by getting incident statuses
            test_url = f"{self.base_url}/incidents/statuses"
            self.root.after(0, lambda: self.log(f"Testing connection with: {test_url}"))

            response = self.session.get(test_url, timeout=30)

            if response.status_code == 200:
                self.root.after(0, self._connection_success)
            elif response.status_code == 401:
                self.root.after(0, lambda: self._connection_failed(
                    "AUTHENTICATION FAILED\n\n"
                    "The username or password is incorrect.\n\n"
                    "Please verify:\n"
                    "• Username is correct (default is 'Administrator')\n"
                    "• Password is correct\n"
                    "• The account is not locked out",
                    is_auth_error=True))
            elif response.status_code == 403:
                self.root.after(0, lambda: self._connection_failed(
                    "ACCESS FORBIDDEN\n\n"
                    "The credentials are valid but this user does not have\n"
                    "permission to access the API.\n\n"
                    "Please verify the user has appropriate role permissions.",
                    is_auth_error=True))
            elif response.status_code == 404:
                self.root.after(0, lambda: self._connection_failed("API endpoint not found. Check server URL and DLP version."))
            else:
                self.root.after(0, lambda: self._connection_failed(f"Connection failed: HTTP {response.status_code} - {response.text[:200]}"))

        except requests.exceptions.SSLError as e:
            error_msg = f"SSL Error: {str(e)}\n\nTry checking if the server URL is correct."
            self.root.after(0, lambda: self._connection_failed(error_msg))
        except requests.exceptions.ConnectionError as e:
            error_str = str(e)
            if "getaddrinfo failed" in error_str or "Name or service not known" in error_str:
                error_msg = "DNS resolution failed. Check the server hostname."
            elif "Connection refused" in error_str:
                error_msg = "Connection refused. Server may be down or port blocked."
            elif "timed out" in error_str.lower():
                error_msg = "Connection timed out. Check network/firewall."
            else:
                error_msg = f"Connection error: {error_str[:200]}"
            self.root.after(0, lambda msg=error_msg: self._connection_failed(msg))
        except requests.exceptions.Timeout:
            self.root.after(0, lambda: self._connection_failed("Connection timeout. Server may be slow or unreachable."))
        except Exception as e:
            self.root.after(0, lambda: self._connection_failed(f"Error: {type(e).__name__}: {str(e)}"))

    def _connection_success(self):
        """Handle successful connection"""
        self.is_connected = True
        self.connect_btn.config(text="Disconnect")
        self.conn_status_label.config(text=f"Connected to {self.server_entry.get()}", foreground="green")
        self.fetch_report_btn.config(state="normal")
        self.load_incidents_btn.config(state="normal")
        self.update_progress(100, "Connected")
        self.log("Successfully connected to Enforce server", "SUCCESS")

        # Store environment URL for this session
        env = self.env_var.get()
        self.log(f"Environment: {self.ENVIRONMENTS[env]['name']}")

    def _on_source_type_change(self):
        """Handle switching between Report and Incident IDs source"""
        source_type = self.source_type_var.get()
        if source_type == "report":
            self.incident_ids_row_frame.grid_forget()
            self.report_row_frame.grid(row=1, column=0, columnspan=7, sticky="ew")
        else:
            self.report_row_frame.grid_forget()
            self.incident_ids_row_frame.grid(row=1, column=0, columnspan=7, sticky="ew")

    def load_incident_ids(self):
        """Load incidents from manually entered IDs"""
        ids_text = self.incident_ids_entry.get().strip()
        if not ids_text:
            messagebox.showerror("Error", "Please enter incident IDs")
            return

        # Parse incident IDs (supports comma-separated and ranges)
        incident_ids = self._parse_incident_ids(ids_text)

        if not incident_ids:
            messagebox.showerror("Error", "No valid incident IDs found.\n\nUse comma-separated values (1,2,3) or ranges (100-200)")
            return

        self.log(f"Loading {len(incident_ids)} incident IDs...")

        # Convert to incident format expected by the rest of the app
        self.current_incidents = [{"incidentId": id} for id in incident_ids]
        count = len(self.current_incidents)

        self.log(f"Loaded {count} incident IDs", "SUCCESS")
        self.incident_count_label.config(text=f"{count} incidents loaded")
        self.fetch_details_btn.config(state="normal")
        self.test_download_btn.config(state="normal")
        self.export_btn.config(state="normal")
        self.update_progress(100, f"Loaded {count} incidents")

    def _parse_incident_ids(self, ids_text):
        """Parse incident IDs from text (comma-separated and/or ranges)"""
        ids = set()

        # Split by comma, semicolon, or whitespace
        parts = [p.strip() for p in ids_text.replace(';', ',').replace('\n', ',').replace('\t', ',').split(',')]

        for part in parts:
            if not part:
                continue

            # Check if it's a range (e.g., 100-200)
            if '-' in part and not part.startswith('-'):
                try:
                    range_parts = part.split('-')
                    if len(range_parts) == 2:
                        start = int(range_parts[0].strip())
                        end = int(range_parts[1].strip())
                        if start <= end and (end - start) <= 10000:  # Limit range size
                            for i in range(start, end + 1):
                                ids.add(i)
                        else:
                            self.log(f"Invalid range: {part} (max 10000 IDs per range)", "WARNING")
                    else:
                        # Might be a negative number, try parsing as single ID
                        ids.add(int(part))
                except ValueError:
                    self.log(f"Invalid ID or range: {part}", "WARNING")
            else:
                # Single ID
                try:
                    ids.add(int(part))
                except ValueError:
                    self.log(f"Invalid ID: {part}", "WARNING")

        return sorted(list(ids))

    def _connection_failed(self, message, is_auth_error=False):
        """Handle failed connection"""
        self.is_connected = False
        self.session = None

        if is_auth_error:
            self.conn_status_label.config(text="Authentication Failed - Check Credentials", foreground="red")
            self.update_progress(0, "Authentication failed")
            # Highlight password field
            self.password_entry.focus_set()
            self.password_entry.selection_range(0, tk.END)
        else:
            self.conn_status_label.config(text="Connection Failed", foreground="red")
            self.update_progress(0, "Connection failed")

        self.log(message, "ERROR")
        messagebox.showerror("Connection Failed", message)

    def disconnect(self):
        """Disconnect from the server"""
        if self.session:
            try:
                # Call logoff endpoint
                logoff_url = f"{self.base_url}/logoff"
                self.session.post(logoff_url, timeout=10)
            except:
                pass
            self.session = None

        self.is_connected = False
        self.connect_btn.config(text="Connect")
        self.conn_status_label.config(text="Not Connected", foreground="red")
        self.fetch_report_btn.config(state="disabled")
        self.load_incidents_btn.config(state="disabled")
        self.fetch_details_btn.config(state="disabled")
        self.export_btn.config(state="disabled")
        self.update_progress(0, "Disconnected")
        self.log("Disconnected from server")

    def fetch_report(self):
        """Fetch saved report by ID"""
        report_id = self.report_id_entry.get().strip()
        if not report_id:
            messagebox.showerror("Error", "Please enter a Report ID")
            return

        if not report_id.isdigit():
            messagebox.showerror("Error", "Report ID must be a number")
            return

        self.log(f"Fetching saved report ID: {report_id}...")
        self.update_progress(0, "Fetching report...")

        # Disable buttons during fetch
        self.fetch_report_btn.config(state="disabled")

        thread = threading.Thread(target=self._fetch_report_thread, args=(report_id,))
        thread.daemon = True
        thread.start()

    def _fetch_report_thread(self, report_id):
        """Thread function for fetching report"""
        try:
            report_url = f"{self.base_url}/savedReport/{report_id}"
            response = self.session.get(report_url, timeout=60)

            if response.status_code == 200:
                report_data = response.json()
                self.root.after(0, lambda: self._process_report(report_data))
            elif response.status_code == 404:
                self.root.after(0, lambda: self._fetch_report_failed(f"Report ID {report_id} not found"))
            else:
                self.root.after(0, lambda: self._fetch_report_failed(f"Failed to fetch report: HTTP {response.status_code}"))

        except Exception as e:
            self.root.after(0, lambda: self._fetch_report_failed(f"Error fetching report: {str(e)}"))

    def _process_report(self, report_data):
        """Process the fetched report data"""
        self.log("Report fetched successfully", "SUCCESS")
        self.log(f"Report data: {json.dumps(report_data, indent=2)[:500]}...")

        # Extract incident IDs from report
        # The report structure may vary, so we need to handle different formats
        incidents = []

        # Try different possible structures
        if isinstance(report_data, dict):
            if "incidents" in report_data:
                incidents = report_data["incidents"]
            elif "incidentIds" in report_data:
                incidents = [{"incidentId": id} for id in report_data["incidentIds"]]
            elif "data" in report_data:
                incidents = report_data["data"]
            elif "incidentId" in report_data:
                incidents = [report_data]
        elif isinstance(report_data, list):
            incidents = report_data

        if incidents:
            self.current_incidents = incidents
            count = len(incidents)
            self.log(f"Found {count} incidents in report")
            self.incident_count_label.config(text=f"{count} incidents loaded")
            self.fetch_details_btn.config(state="normal")
            self.test_download_btn.config(state="normal")
            self.export_btn.config(state="normal")
            self.update_progress(100, f"Loaded {count} incidents")
        else:
            self.log("No incidents found in report. Attempting to list incidents...", "WARNING")
            # Try listing incidents using the report filters
            self._list_incidents_from_report(report_data)

        self.fetch_report_btn.config(state="normal")

    def _list_incidents_from_report(self, report_data):
        """List incidents based on report configuration with pagination"""
        try:
            # Use select fields from the saved report if available
            select_fields = report_data.get("select", [
                {"name": "incidentId"},
                {"name": "policyName"},
                {"name": "matchCount"},
                {"name": "severityId"},
                {"name": "incidentStatusName"},
                {"name": "messageDate"}
            ])

            # Make sure incidentId is always included
            has_incident_id = any(
                (isinstance(f, dict) and f.get("name") == "incidentId") or f == "incidentId"
                for f in select_fields
            )
            if not has_incident_id:
                select_fields.insert(0, {"name": "incidentId"})

            self.log(f"Querying incidents with report filters...")

            all_incidents = []
            page_number = 1
            page_size = 1000  # Larger page size for efficiency
            total_available = None

            incidents_url = f"{self.base_url}/incidents"

            while True:
                # Build request body matching API format
                request_body = {
                    "select": select_fields,
                    "page": {
                        "type": "offset",
                        "pageNumber": page_number,
                        "pageSize": page_size
                    }
                }

                # Use filter from report if available
                if isinstance(report_data, dict) and "filter" in report_data:
                    request_body["filter"] = report_data["filter"]

                # Use orderBy from report if available
                if isinstance(report_data, dict) and "orderBy" in report_data:
                    request_body["orderBy"] = report_data["orderBy"]

                self.log(f"Fetching page {page_number}...")
                self.update_progress(0, f"Fetching incidents page {page_number}...")

                response = self.session.post(incidents_url, json=request_body, timeout=120)

                if response.status_code == 200:
                    result = response.json()

                    # Handle different response formats
                    incidents = []
                    if isinstance(result, dict):
                        incidents = result.get("incidents", result.get("data", []))
                        if total_available is None:
                            total_available = result.get("total", result.get("totalCount", 0))
                    elif isinstance(result, list):
                        incidents = result

                    if incidents:
                        all_incidents.extend(incidents)
                        self.log(f"Page {page_number}: Got {len(incidents)} incidents (total so far: {len(all_incidents)})")

                        # Check if we need more pages
                        if len(incidents) < page_size:
                            break
                        if total_available and len(all_incidents) >= total_available:
                            break

                        page_number += 1
                        if total_available:
                            progress = (len(all_incidents) / total_available) * 100
                            self.update_progress(progress, f"Loaded {len(all_incidents)}/{total_available} incidents")
                    else:
                        break
                else:
                    error_detail = ""
                    try:
                        error_detail = response.json()
                    except:
                        error_detail = response.text[:500]
                    self.log(f"Failed to list incidents: HTTP {response.status_code}", "ERROR")
                    self.log(f"Error details: {error_detail}", "ERROR")
                    break

            if all_incidents:
                self.current_incidents = all_incidents
                count = len(self.current_incidents)
                self.log(f"Listed {count} incidents (total available: {total_available})", "SUCCESS")
                self.incident_count_label.config(text=f"{count} incidents loaded")
                self.fetch_details_btn.config(state="normal")
                self.test_download_btn.config(state="normal")
                self.export_btn.config(state="normal")
                self.update_progress(100, f"Loaded {count} incidents")
            else:
                self.log("No incidents found matching report criteria", "WARNING")

        except Exception as e:
            self.log(f"Error listing incidents: {str(e)}", "ERROR")

    def _fetch_report_failed(self, message):
        """Handle failed report fetch"""
        self.fetch_report_btn.config(state="normal")
        self.update_progress(0, "Failed")
        self.log(message, "ERROR")
        messagebox.showerror("Error", message)

    def test_download(self):
        """Download a limited number of incidents for testing"""
        if not self.current_incidents:
            messagebox.showinfo("Info", "No incidents loaded. Please fetch a report first.")
            return

        total = len(self.current_incidents)

        # Ask how many to download
        test_count = simpledialog.askinteger(
            "Test Download",
            f"Total incidents available: {total}\n\n"
            f"How many incidents to download for testing?\n"
            f"(Recommended: 5-20 for quick verification)",
            initialvalue=min(10, total),
            minvalue=1,
            maxvalue=total,
            parent=self.root
        )

        if test_count is None:
            return

        self.log(f"Test download: fetching {test_count} of {total} incidents...")

        # Store the limited set of incidents to fetch
        self._test_incidents = self.current_incidents[:test_count]

        # Use single thread for test (simpler)
        self._thread_count = min(5, test_count)
        self._stop_fetch = False
        self._fetch_in_progress = True

        self.fetch_details_btn.config(state="disabled")
        self.test_download_btn.config(state="disabled")
        self.fetch_report_btn.config(state="disabled")
        self.export_btn.config(state="disabled")
        self.stop_btn.config(state="normal")
        self.resume_btn.config(state="disabled")

        thread = threading.Thread(target=self._fetch_test_details_thread, args=(test_count,))
        thread.daemon = True
        thread.start()

    def _fetch_test_details_thread(self, test_count):
        """Thread function for fetching test incident details"""
        thread_count = getattr(self, '_thread_count', 5)
        incidents_to_fetch = self._test_incidents

        # Initialize cache directory for storing incident data
        self._init_cache_dir()
        self._fetched_incident_ids = set()

        completed = [0]
        lock = threading.Lock()
        start_time = datetime.now()

        def fetch_single_incident(incident):
            if self._stop_fetch:
                return None, None
            incident_id = incident.get("incidentId") or incident.get("id")
            if not incident_id:
                return None, None
            details = self._fetch_incident_details(incident_id)
            return incident_id, details

        def update_progress_callback(incident_id, details):
            with lock:
                completed[0] += 1
                progress = (completed[0] / test_count) * 100

                self.root.after(0, lambda p=progress, c=completed[0], t=test_count:
                               self.update_progress(p, f"Test: {c}/{t} incidents fetched"))

                # Save to cache file instead of keeping in memory
                if details:
                    self._save_incident_to_cache(incident_id, details)

        self.root.after(0, lambda: self.log(f"Starting test download ({test_count} incidents, {thread_count} threads)..."))
        self.logger.info(f"Starting test download of {test_count} incidents")

        with ThreadPoolExecutor(max_workers=thread_count) as executor:
            future_to_incident = {
                executor.submit(fetch_single_incident, incident): incident
                for incident in incidents_to_fetch
            }

            for future in as_completed(future_to_incident):
                if self._stop_fetch:
                    executor.shutdown(wait=False, cancel_futures=True)
                    break
                try:
                    incident_id, details = future.result()
                    if incident_id:
                        update_progress_callback(incident_id, details)
                except Exception as e:
                    self.root.after(0, lambda err=str(e): self.log(f"Error: {err}", "ERROR"))

        elapsed = (datetime.now() - start_time).total_seconds()
        count = len(self._fetched_incident_ids)

        if self._stop_fetch:
            self.root.after(0, lambda: self.log(f"Test stopped. Fetched {count} incidents.", "WARNING"))
        else:
            self.root.after(0, lambda: self.log(
                f"Test complete: {count} incidents in {elapsed:.1f}s", "SUCCESS"))
            self.root.after(0, lambda: self.log(
                f"Review the data, then use 'Fetch All' for the complete download.", "INFO"))

        self.root.after(0, self._test_download_complete)

    def _test_download_complete(self):
        """Handle completion of test download"""
        self._fetch_in_progress = False
        count = len(self._fetched_incident_ids)
        self.update_progress(100, f"Test: {count} incidents fetched")
        self.fetch_details_btn.config(state="normal")
        self.test_download_btn.config(state="normal")
        self.fetch_report_btn.config(state="normal")
        self.export_btn.config(state="normal")
        self.stop_btn.config(state="disabled")
        self.incident_count_label.config(text=f"TEST: {count} incidents with full details")

    def fetch_all_incident_details(self):
        """Fetch detailed information for all incidents"""
        if not self.current_incidents:
            messagebox.showinfo("Info", "No incidents loaded. Please fetch a report first.")
            return

        # Ask for number of threads
        total = len(self.current_incidents)
        default_threads = min(10, max(1, total // 100))  # Scale threads based on count
        if total > 100:
            thread_count = simpledialog.askinteger(
                "Thread Count",
                f"Found {total} incidents.\n\nEnter number of parallel threads (1-20):\n(More threads = faster but more server load)",
                initialvalue=default_threads,
                minvalue=1,
                maxvalue=20,
                parent=self.root
            )
            if thread_count is None:
                return
        else:
            thread_count = default_threads

        self.log(f"Starting to fetch details for {total} incidents using {thread_count} threads...")
        self.fetch_details_btn.config(state="disabled")
        self.test_download_btn.config(state="disabled")
        self.fetch_report_btn.config(state="disabled")
        self.export_btn.config(state="disabled")
        self.stop_btn.config(state="normal")
        self.resume_btn.config(state="disabled")

        # Store thread count for the worker
        self._thread_count = thread_count
        self._stop_fetch = False
        self._fetch_in_progress = True

        thread = threading.Thread(target=self._fetch_all_details_thread)
        thread.daemon = True
        thread.start()

    def _fetch_all_details_thread(self):
        """Thread function for fetching all incident details with thread pool"""
        total = len(self.current_incidents)
        thread_count = getattr(self, '_thread_count', 5)

        # Initialize cache directory for storing incident data
        self._init_cache_dir()
        self._fetched_incident_ids = set()

        # Track progress
        completed = [0]  # Use list to allow modification in nested function
        lock = threading.Lock()
        start_time = datetime.now()

        def fetch_single_incident(incident):
            """Fetch details for a single incident (called by thread pool)"""
            if self._stop_fetch:
                return None, None
            incident_id = incident.get("incidentId") or incident.get("id")
            if not incident_id:
                return None, None

            details = self._fetch_incident_details(incident_id)
            return incident_id, details

        def update_progress_callback(incident_id, details):
            """Update progress after each incident is fetched"""
            with lock:
                completed[0] += 1
                progress = (completed[0] / total) * 100

                # Calculate ETA
                elapsed = (datetime.now() - start_time).total_seconds()
                if completed[0] > 0:
                    rate = completed[0] / elapsed
                    remaining = total - completed[0]
                    eta_seconds = remaining / rate if rate > 0 else 0
                    eta_str = f"ETA: {int(eta_seconds // 60)}m {int(eta_seconds % 60)}s"
                else:
                    eta_str = "Calculating..."

                self.root.after(0, lambda p=progress, c=completed[0], t=total, e=eta_str:
                               self.update_progress(p, f"Fetched {c}/{t} incidents - {e}"))

                # Save to cache file instead of keeping in memory
                if details:
                    self._save_incident_to_cache(incident_id, details)

                # Log every 10 incidents or for small batches
                if completed[0] % max(1, total // 20) == 0 or completed[0] == total:
                    self.root.after(0, lambda c=completed[0], t=total:
                                   self.log(f"Progress: {c}/{t} incidents fetched"))

        # Use ThreadPoolExecutor for parallel fetching
        self.root.after(0, lambda: self.log(f"Starting parallel fetch with {thread_count} threads..."))
        self.logger.info(f"Starting fetch of {total} incidents with {thread_count} threads")

        with ThreadPoolExecutor(max_workers=thread_count) as executor:
            # Submit all tasks
            future_to_incident = {
                executor.submit(fetch_single_incident, incident): incident
                for incident in self.current_incidents
            }

            # Process results as they complete
            for future in as_completed(future_to_incident):
                if self._stop_fetch:
                    executor.shutdown(wait=False, cancel_futures=True)
                    break
                try:
                    incident_id, details = future.result()
                    if incident_id:
                        update_progress_callback(incident_id, details)
                except Exception as e:
                    self.root.after(0, lambda err=str(e): self.log(f"Error fetching incident: {err}", "ERROR"))

        # Calculate final stats
        elapsed = (datetime.now() - start_time).total_seconds()
        rate = completed[0] / elapsed if elapsed > 0 else 0

        if self._stop_fetch:
            self.root.after(0, lambda: self._fetch_all_details_complete(was_stopped=True))
        else:
            self.root.after(0, lambda: self.log(
                f"Fetch complete: {completed[0]} incidents in {elapsed:.1f}s ({rate:.1f} incidents/sec)", "SUCCESS"))
            self.root.after(0, lambda: self._fetch_all_details_complete(was_stopped=False))

    def _fetch_incident_details(self, incident_id):
        """Fetch all details for a single incident"""
        details = {
            "incidentId": incident_id,
            "fetchedAt": datetime.now().isoformat()
        }

        # List of detail endpoints to call (including internal APIs for full details)
        endpoints = [
            ("staticAttributes", f"/incidents/{incident_id}/staticAttributes"),
            ("editableAttributes", f"/incidents/{incident_id}/editableAttributes"),
            ("components", f"/incidents/{incident_id}/components"),
            ("matches", f"/incidents/{incident_id}/components/matches"),
            # Internal API provides more detailed component/match info with violations
            ("componentMatches", f"/internal/incidents/{incident_id}/components/matches"),
            ("correlations", f"/incidents/{incident_id}/correlations"),
            ("history", f"/incidents/{incident_id}/history"),
            ("policyMatches", f"/incidents/{incident_id}/policymatches/"),
            ("incidentMetadata", f"/internal/incidents/incidentDetailsMetadata"),
        ]

        for name, endpoint in endpoints:
            try:
                url = f"{self.base_url}{endpoint}"
                response = self.session.get(url, timeout=30)

                if response.status_code == 200:
                    details[name] = response.json()
                elif response.status_code == 404:
                    details[name] = None  # Data not available for this policy type
                else:
                    details[name] = {"error": f"HTTP {response.status_code}"}

            except Exception as e:
                details[name] = {"error": str(e)}

        # Collect ALL component IDs from multiple sources
        import base64
        all_component_ids = set()

        # Source 1: components list
        if details.get("components") and isinstance(details["components"], list):
            for component in details["components"]:
                comp_id = component.get("componentId") or component.get("id")
                if comp_id:
                    all_component_ids.add(str(comp_id))

        # Source 2: componentMatches (internal API)
        if details.get("componentMatches") and isinstance(details["componentMatches"], list):
            for comp_match in details["componentMatches"]:
                comp_id = comp_match.get("messageComponentId") or comp_match.get("componentId") or comp_match.get("id")
                if comp_id:
                    all_component_ids.add(str(comp_id))

        # Source 3: matches list
        if details.get("matches") and isinstance(details["matches"], list):
            for match in details["matches"]:
                comp_id = match.get("componentId") or match.get("messageComponentId")
                if comp_id:
                    all_component_ids.add(str(comp_id))

        # Source 4: staticAttributes.infoMap.attachmentInfo
        if details.get("staticAttributes") and isinstance(details["staticAttributes"], dict):
            info_map = details["staticAttributes"].get("infoMap", {})
            if isinstance(info_map, dict):
                attachment_info = info_map.get("attachmentInfo", [])
                if isinstance(attachment_info, list):
                    for att in attachment_info:
                        if isinstance(att, dict):
                            comp_id = att.get("componentId") or att.get("id")
                            if comp_id:
                                all_component_ids.add(str(comp_id))

        # Fetch individual component data for ALL components found
        details["componentData"] = {}
        details["componentCount"] = len(all_component_ids)

        for comp_id in all_component_ids:
            try:
                url = f"{self.base_url}/incidents/{incident_id}/components/{comp_id}"
                response = self.session.get(url, timeout=60)
                if response.status_code == 200:
                    content_type = response.headers.get('Content-Type', '')

                    # Extract filename from Content-Disposition header if present
                    original_filename = None
                    content_disp = response.headers.get('Content-Disposition', '')
                    if content_disp:
                        # Parse filename from Content-Disposition header
                        # Format: attachment; filename="example.pdf" or filename*=UTF-8''example.pdf
                        import re
                        # Try quoted filename first
                        match = re.search(r'filename[*]?=["\']?([^"\';\r\n]+)["\']?', content_disp, re.IGNORECASE)
                        if match:
                            original_filename = match.group(1).strip()
                            # Handle UTF-8 encoded filenames
                            if original_filename.startswith("UTF-8''"):
                                from urllib.parse import unquote
                                original_filename = unquote(original_filename[7:])

                    if 'application/json' in content_type:
                        details["componentData"][comp_id] = response.json()
                    else:
                        # Binary component data - store as base64
                        details["componentData"][comp_id] = {
                            "contentType": content_type,
                            "size": len(response.content),
                            "originalFilename": original_filename,
                            "dataBase64": base64.b64encode(response.content).decode('utf-8')
                        }
                else:
                    details["componentData"][comp_id] = {"error": f"HTTP {response.status_code}"}
            except Exception as e:
                details["componentData"][comp_id] = {"error": str(e)}

        # Try to get message body (may not be available for all incident types)
        try:
            url = f"{self.base_url}/incidents/{incident_id}/messageBody"
            response = self.session.get(url, timeout=30)
            if response.status_code == 200:
                details["messageBody"] = response.json()
        except:
            pass

        # Try to get original message
        try:
            url = f"{self.base_url}/incidents/{incident_id}/originalMessage"
            response = self.session.get(url, timeout=60)  # Longer timeout for attachments
            if response.status_code == 200:
                content_type = response.headers.get('Content-Type', '')
                if 'application/json' in content_type:
                    details["originalMessage"] = response.json()
                else:
                    # Binary data - store as base64
                    import base64
                    details["originalMessage"] = {
                        "contentType": content_type,
                        "size": len(response.content),
                        "dataBase64": base64.b64encode(response.content).decode('utf-8')
                    }
        except Exception as e:
            details["originalMessage"] = {"error": str(e)}

        return details

    def _fetch_all_details_complete(self, was_stopped=False):
        """Handle completion of fetching all details"""
        self._fetch_in_progress = False
        count = len(self._fetched_incident_ids)

        if was_stopped:
            self.log(f"Fetch stopped. Got {count} incidents before stopping.", "WARNING")
            self.update_progress(0, f"Stopped - {count} incidents fetched")
            # Save state for resume
            self._save_resume_state()
            self.logger.warning(f"Fetch stopped with {count} incidents cached")
        else:
            self.log(f"Completed fetching details for {count} incidents", "SUCCESS")
            self.update_progress(100, f"Fetched details for {count} incidents")
            # Clear resume state on successful completion
            self._clear_resume_state()
            self.logger.info(f"Fetch completed successfully: {count} incidents cached")

        self.fetch_details_btn.config(state="normal")
        self.test_download_btn.config(state="normal")
        self.fetch_report_btn.config(state="normal")
        self.export_btn.config(state="normal")
        self.stop_btn.config(state="disabled")
        self._update_resume_button()
        self.incident_count_label.config(text=f"{count} incidents with full details")

    def stop_fetch(self):
        """Stop the current fetch operation"""
        if self._fetch_in_progress:
            self._stop_fetch = True
            self.log("Stopping fetch... please wait for current operations to complete.", "WARNING")
            self.stop_btn.config(state="disabled")
            self.update_progress(0, "Stopping...")

    def _save_resume_state(self):
        """Save current state for resume capability"""
        try:
            state = {
                "server": self.server_entry.get(),
                "report_id": self.report_id_entry.get(),
                "env_type": self.env_var.get(),
                "current_incidents": self.current_incidents,
                "fetched_ids": list(self.incident_details.keys()),
                "incident_details": {str(k): v for k, v in self.incident_details.items()},
                "saved_at": datetime.now().isoformat()
            }
            with open(self._resume_file, 'w', encoding='utf-8') as f:
                json.dump(state, f, indent=2, default=str)
            self.log(f"Saved resume state: {len(self.incident_details)} incidents fetched", "SUCCESS")
        except Exception as e:
            self.log(f"Failed to save resume state: {str(e)}", "ERROR")

    def _load_resume_state(self):
        """Load saved resume state"""
        try:
            if os.path.exists(self._resume_file):
                with open(self._resume_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except Exception as e:
            self.log(f"Failed to load resume state: {str(e)}", "ERROR")
        return None

    def _clear_resume_state(self):
        """Clear the resume state file"""
        try:
            if os.path.exists(self._resume_file):
                os.remove(self._resume_file)
                self.log("Cleared resume state file")
        except Exception as e:
            self.log(f"Failed to clear resume state: {str(e)}", "ERROR")

    def _update_resume_button(self):
        """Update resume button state based on saved state"""
        state = self._load_resume_state()
        if state and state.get("fetched_ids"):
            fetched = len(state.get("fetched_ids", []))
            total = len(state.get("current_incidents", []))
            remaining = total - fetched
            if remaining > 0:
                self.resume_btn.config(state="normal")
                self.log(f"Resume available: {fetched}/{total} incidents fetched, {remaining} remaining")
            else:
                self.resume_btn.config(state="disabled")
        else:
            self.resume_btn.config(state="disabled")

    def resume_fetch(self):
        """Resume a previously interrupted fetch"""
        state = self._load_resume_state()
        if not state:
            messagebox.showinfo("Info", "No saved state to resume from.")
            return

        # Check if connected to same server
        current_server = self.server_entry.get().strip()
        saved_server = state.get("server", "")

        if not self.is_connected:
            messagebox.showerror("Error", "Please connect to the server first before resuming.")
            return

        if current_server != saved_server:
            if not messagebox.askyesno("Different Server",
                                       f"Saved state is from a different server:\n"
                                       f"Saved: {saved_server}\n"
                                       f"Current: {current_server}\n\n"
                                       f"Continue anyway?"):
                return

        # Restore state
        self.current_incidents = state.get("current_incidents", [])
        self.incident_details = {
            # Convert string keys back to int if they were int
            (int(k) if k.isdigit() else k): v
            for k, v in state.get("incident_details", {}).items()
        }
        fetched_ids = set(state.get("fetched_ids", []))

        # Filter to only unfetched incidents
        remaining_incidents = [
            inc for inc in self.current_incidents
            if (inc.get("incidentId") or inc.get("id")) not in fetched_ids
            and str(inc.get("incidentId") or inc.get("id")) not in fetched_ids
        ]

        if not remaining_incidents:
            messagebox.showinfo("Info", "All incidents have already been fetched.")
            self._clear_resume_state()
            self._update_resume_button()
            return

        total_fetched = len(fetched_ids)
        total_remaining = len(remaining_incidents)
        total = len(self.current_incidents)

        self.log(f"Resuming fetch: {total_fetched} already fetched, {total_remaining} remaining", "INFO")

        # Ask for thread count
        default_threads = min(10, max(1, total_remaining // 100))
        if total_remaining > 100:
            thread_count = simpledialog.askinteger(
                "Thread Count",
                f"Resuming: {total_fetched} fetched, {total_remaining} remaining.\n\n"
                f"Enter number of parallel threads (1-20):",
                initialvalue=default_threads,
                minvalue=1,
                maxvalue=20,
                parent=self.root
            )
            if thread_count is None:
                return
        else:
            thread_count = default_threads

        self._thread_count = thread_count
        self._remaining_incidents = remaining_incidents
        self._stop_fetch = False
        self._fetch_in_progress = True

        self.fetch_details_btn.config(state="disabled")
        self.fetch_report_btn.config(state="disabled")
        self.export_btn.config(state="disabled")
        self.stop_btn.config(state="normal")
        self.resume_btn.config(state="disabled")

        self.incident_count_label.config(text=f"{total} incidents ({total_fetched} fetched, {total_remaining} remaining)")

        thread = threading.Thread(target=self._fetch_remaining_details_thread)
        thread.daemon = True
        thread.start()

    def _fetch_remaining_details_thread(self):
        """Thread function for fetching remaining incident details"""
        remaining = self._remaining_incidents
        total_remaining = len(remaining)
        already_fetched = len(self.incident_details)
        total = already_fetched + total_remaining
        thread_count = getattr(self, '_thread_count', 5)

        completed = [0]
        lock = threading.Lock()
        start_time = datetime.now()

        def fetch_single_incident(incident):
            if self._stop_fetch:
                return None, None
            incident_id = incident.get("incidentId") or incident.get("id")
            if not incident_id:
                return None, None
            details = self._fetch_incident_details(incident_id)
            return incident_id, details

        def update_progress_callback(incident_id, details):
            with lock:
                completed[0] += 1
                current_total = already_fetched + completed[0]
                progress = (current_total / total) * 100

                elapsed = (datetime.now() - start_time).total_seconds()
                if completed[0] > 0:
                    rate = completed[0] / elapsed
                    remaining_count = total_remaining - completed[0]
                    eta_seconds = remaining_count / rate if rate > 0 else 0
                    eta_str = f"ETA: {int(eta_seconds // 60)}m {int(eta_seconds % 60)}s"
                else:
                    eta_str = "Calculating..."

                self.root.after(0, lambda p=progress, c=current_total, t=total, e=eta_str:
                               self.update_progress(p, f"Fetched {c}/{t} incidents - {e}"))

                if details:
                    self.incident_details[incident_id] = details

                if completed[0] % max(1, total_remaining // 20) == 0 or completed[0] == total_remaining:
                    self.root.after(0, lambda c=completed[0], t=total_remaining:
                                   self.log(f"Resume progress: {c}/{t} remaining incidents fetched"))

        self.root.after(0, lambda: self.log(f"Resuming parallel fetch with {thread_count} threads..."))

        with ThreadPoolExecutor(max_workers=thread_count) as executor:
            future_to_incident = {
                executor.submit(fetch_single_incident, incident): incident
                for incident in remaining
            }

            for future in as_completed(future_to_incident):
                if self._stop_fetch:
                    executor.shutdown(wait=False, cancel_futures=True)
                    break
                try:
                    incident_id, details = future.result()
                    if incident_id:
                        update_progress_callback(incident_id, details)
                except Exception as e:
                    self.root.after(0, lambda err=str(e): self.log(f"Error fetching incident: {err}", "ERROR"))

        elapsed = (datetime.now() - start_time).total_seconds()
        rate = completed[0] / elapsed if elapsed > 0 else 0

        if self._stop_fetch:
            self.root.after(0, lambda: self._fetch_all_details_complete(was_stopped=True))
        else:
            self.root.after(0, lambda: self.log(
                f"Resume complete: {completed[0]} incidents in {elapsed:.1f}s ({rate:.1f} incidents/sec)", "SUCCESS"))
            self.root.after(0, lambda: self._fetch_all_details_complete(was_stopped=False))

    def export_incidents(self):
        """Export incidents based on selected format"""
        export_format = self.export_format_var.get()
        cached_count = len(self._fetched_incident_ids) if self._fetched_incident_ids else 0
        self.log(f"Export requested - Format: {export_format}")
        self.log(f"Current incidents: {len(self.current_incidents)}, Cached details: {cached_count}")
        self.logger.info(f"Export started: format={export_format}, cached={cached_count}")

        if not cached_count and not self.current_incidents:
            messagebox.showinfo("Info", "No incident data to export")
            return

        # Bring window to front before showing dialog
        self.root.lift()
        self.root.focus_force()

        if export_format == "JSON + Attachments":
            self._export_json_with_attachments()
        elif export_format == "JSON (AI-Friendly)":
            self._export_json_ai_friendly()
        elif export_format == "CSV (Summary)":
            self._export_csv()
        elif export_format == "Individual JSON Files":
            self._export_individual_json()

    def _export_json_with_attachments(self):
        """Export incidents to folder with JSON and attachment files"""
        folder_path = self.output_path_entry.get().strip()

        if not folder_path:
            messagebox.showerror("Error", "Please specify an output folder path")
            return

        self.log(f"Exporting JSON with attachments to: {folder_path}")

        try:
            import base64

            # Create main export folder
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            export_folder = os.path.join(folder_path, f"dlp_export_{timestamp}")
            os.makedirs(export_folder, exist_ok=True)

            # Use cached incidents (streaming from disk) or fall back to current_incidents
            use_cache = bool(self._fetched_incident_ids)
            total_incidents = len(self._fetched_incident_ids) if use_cache else len(self.current_incidents)

            self.log(f"Processing {total_incidents} incidents (from {'cache' if use_cache else 'memory'})...")
            self.logger.info(f"Export processing {total_incidents} incidents from {'cache' if use_cache else 'memory'}")
            attachment_count = 0
            processed_count = 0

            # Create iterator based on data source
            if use_cache:
                incident_iterator = self._iter_cached_incidents()
            else:
                incident_iterator = ((inc.get("incidentId") or inc.get("id") or i, inc)
                                    for i, inc in enumerate(self.current_incidents))

            for incident_id, details in incident_iterator:
                processed_count += 1
                if processed_count % 100 == 0:
                    self.log(f"  Exported {processed_count}/{total_incidents} incidents...")
                    self.root.update_idletasks()
                # Create incident folder
                incident_folder = os.path.join(export_folder, f"incident_{incident_id}")
                os.makedirs(incident_folder, exist_ok=True)

                # Copy details for JSON (without base64 data)
                json_details = {}
                for key, value in details.items():
                    json_details[key] = value

                # Build a map of component IDs to names from staticAttributes.infoMap.attachmentInfo
                component_names = {}
                if "staticAttributes" in details and isinstance(details["staticAttributes"], dict):
                    info_map = details["staticAttributes"].get("infoMap", {})
                    if isinstance(info_map, dict):
                        attachment_info = info_map.get("attachmentInfo", [])
                        if isinstance(attachment_info, list):
                            for att in attachment_info:
                                if isinstance(att, dict):
                                    comp_id = att.get("componentId") or att.get("id")
                                    att_name = att.get("attachmentName") or att.get("name") or att.get("documentName")
                                    if comp_id and att_name:
                                        component_names[str(comp_id)] = att_name

                # Also check components list for names
                if "components" in details and isinstance(details["components"], list):
                    for comp in details["components"]:
                        if isinstance(comp, dict):
                            comp_id = comp.get("componentId") or comp.get("id")
                            comp_name = comp.get("name") or comp.get("documentName") or comp.get("fileName")
                            if comp_id and comp_name and str(comp_id) not in component_names:
                                component_names[str(comp_id)] = comp_name

                # Save original message/email body as file if it's binary
                if "originalMessage" in details and isinstance(details["originalMessage"], dict):
                    orig_msg = details["originalMessage"]
                    if "dataBase64" in orig_msg:
                        # Decode first to detect file type
                        binary_data = base64.b64decode(orig_msg["dataBase64"])

                        # Use magic byte detection for accurate file type
                        detected_ext = detect_file_type_by_magic(binary_data)

                        # Fall back to content type if magic bytes don't detect
                        content_type = orig_msg.get("contentType", "application/octet-stream")
                        ext = detected_ext or self._get_extension_from_content_type(content_type)

                        attachment_filename = f"email_body{ext}"
                        attachment_path = os.path.join(incident_folder, attachment_filename)

                        with open(attachment_path, 'wb') as f:
                            f.write(binary_data)

                        # Update JSON to reference file instead of base64
                        json_details["originalMessage"] = {
                            "contentType": content_type,
                            "detectedType": detected_ext,
                            "size": orig_msg.get("size"),
                            "savedToFile": attachment_filename
                        }
                        attachment_count += 1
                        self.log(f"  Incident {incident_id}: Saved email body: {attachment_filename}{' [detected: ' + detected_ext + ']' if detected_ext else ''}")

                # Save component data (attachments) as files if binary
                if "componentData" in details and isinstance(details["componentData"], dict):
                    json_details["componentData"] = {}
                    attachments_folder = os.path.join(incident_folder, "attachments")

                    for comp_id, comp_data in details["componentData"].items():
                        if isinstance(comp_data, dict) and "dataBase64" in comp_data:
                            # Create attachments subfolder only if we have attachments
                            if not os.path.exists(attachments_folder):
                                os.makedirs(attachments_folder, exist_ok=True)

                            content_type = comp_data.get("contentType", "application/octet-stream")
                            binary_data = base64.b64decode(comp_data["dataBase64"])

                            # Detect actual file type using magic bytes
                            detected_ext = detect_file_type_by_magic(binary_data)

                            # Try to get the original filename - priority order:
                            # 1. From Content-Disposition header (originalFilename)
                            # 2. From attachmentInfo (component_names)
                            # 3. Fall back to content-type based extension
                            original_name = comp_data.get("originalFilename") or component_names.get(str(comp_id))
                            if original_name:
                                # Sanitize filename but preserve the extension
                                safe_name = "".join(c for c in original_name if c.isalnum() or c in '._- ')
                                comp_filename = safe_name if safe_name else f"attachment_{comp_id}"

                                # Check if extension matches magic bytes, correct if needed
                                if '.' in comp_filename:
                                    name_part, current_ext = os.path.splitext(comp_filename)
                                    # If magic bytes detected a different type, use that instead
                                    if detected_ext and current_ext.lower() != detected_ext.lower():
                                        self.log(f"  Incident {incident_id}: Correcting extension {current_ext} -> {detected_ext} (magic byte detection)")
                                        comp_filename = f"{name_part}{detected_ext}"
                                else:
                                    # No extension, use magic byte detection or content-type
                                    ext = detected_ext or self._get_extension_from_content_type(content_type)
                                    comp_filename = f"{comp_filename}{ext}"
                            else:
                                # No original name - use magic bytes first, then content-type
                                ext = detected_ext or self._get_extension_from_content_type(content_type)
                                comp_filename = f"attachment_{comp_id}{ext}"

                            comp_path = os.path.join(attachments_folder, comp_filename)

                            # Handle duplicate filenames
                            if os.path.exists(comp_path):
                                base, ext = os.path.splitext(comp_filename)
                                comp_filename = f"{base}_{comp_id}{ext}"
                                comp_path = os.path.join(attachments_folder, comp_filename)
                            with open(comp_path, 'wb') as f:
                                f.write(binary_data)

                            json_details["componentData"][comp_id] = {
                                "contentType": content_type,
                                "detectedType": detected_ext,
                                "size": len(binary_data),
                                "originalName": original_name,
                                "savedToFile": f"attachments/{comp_filename}"
                            }
                            attachment_count += 1
                            self.log(f"  Incident {incident_id}: Saved attachment: {comp_filename} ({len(binary_data):,} bytes){' [detected: ' + detected_ext + ']' if detected_ext else ''}")
                        else:
                            json_details["componentData"][comp_id] = comp_data

                # Save incident JSON
                json_path = os.path.join(incident_folder, "incident_details.json")
                with open(json_path, 'w', encoding='utf-8') as f:
                    json.dump(json_details, f, indent=2, default=str)

            # Save summary JSON
            summary = {
                "exportDate": datetime.now().isoformat(),
                "environment": self.env_var.get(),
                "server": self.server_entry.get(),
                "totalIncidents": processed_count,
                "totalAttachments": attachment_count,
                "incidentIds": list(self._fetched_incident_ids) if use_cache else [inc.get("incidentId") or inc.get("id") for inc in self.current_incidents]
            }
            summary_path = os.path.join(export_folder, "export_summary.json")
            with open(summary_path, 'w', encoding='utf-8') as f:
                json.dump(summary, f, indent=2, default=str)

            self.log(f"Export complete: {processed_count} incidents, {attachment_count} attachments", "SUCCESS")
            self.logger.info(f"Export complete: {processed_count} incidents, {attachment_count} attachments to {export_folder}")
            messagebox.showinfo("Export Complete",
                              f"Exported to: {export_folder}\n\n"
                              f"Incidents: {processed_count}\n"
                              f"Attachments: {attachment_count}")

        except Exception as e:
            self.log(f"Export failed: {str(e)}", "ERROR")
            import traceback
            self.log(f"Traceback: {traceback.format_exc()}", "ERROR")
            messagebox.showerror("Export Failed", str(e))

    def _get_extension_from_content_type(self, content_type):
        """Get file extension from content type"""
        content_type = content_type.lower().split(';')[0].strip()
        extensions = {
            'application/pdf': '.pdf',
            'application/msword': '.doc',
            'application/vnd.openxmlformats-officedocument.wordprocessingml.document': '.docx',
            'application/vnd.ms-excel': '.xls',
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': '.xlsx',
            'application/vnd.ms-powerpoint': '.ppt',
            'application/vnd.openxmlformats-officedocument.presentationml.presentation': '.pptx',
            'application/zip': '.zip',
            'application/x-zip-compressed': '.zip',
            'application/x-rar-compressed': '.rar',
            'application/x-7z-compressed': '.7z',
            'application/gzip': '.gz',
            'application/json': '.json',
            'application/xml': '.xml',
            'text/plain': '.txt',
            'text/html': '.html',
            'text/csv': '.csv',
            'text/xml': '.xml',
            'image/jpeg': '.jpg',
            'image/png': '.png',
            'image/gif': '.gif',
            'image/bmp': '.bmp',
            'image/tiff': '.tiff',
            'message/rfc822': '.eml',
            'application/octet-stream': '.bin',
        }
        return extensions.get(content_type, '.bin')

    def _extract_text_from_binary(self, binary_data, content_type, filename=None):
        """Extract readable text from binary document data"""
        import base64

        text_content = None
        extraction_method = None

        try:
            content_type_lower = content_type.lower() if content_type else ""

            # Plain text files
            if 'text/' in content_type_lower:
                try:
                    text_content = binary_data.decode('utf-8')
                    extraction_method = "text/utf-8"
                except UnicodeDecodeError:
                    try:
                        text_content = binary_data.decode('latin-1')
                        extraction_method = "text/latin-1"
                    except:
                        pass

            # PDF files
            elif 'pdf' in content_type_lower or (filename and filename.lower().endswith('.pdf')):
                if HAS_PYPDF2:
                    try:
                        pdf_file = io.BytesIO(binary_data)
                        pdf_reader = PyPDF2.PdfReader(pdf_file)
                        text_parts = []
                        for page in pdf_reader.pages:
                            page_text = page.extract_text()
                            if page_text:
                                text_parts.append(page_text)
                        if text_parts:
                            text_content = "\n\n".join(text_parts)
                            extraction_method = "PyPDF2"
                    except Exception as e:
                        text_content = f"[PDF extraction failed: {str(e)}]"
                else:
                    text_content = "[PDF extraction requires PyPDF2 - run: pip install PyPDF2]"

            # Word documents (.docx)
            elif 'wordprocessingml' in content_type_lower or (filename and filename.lower().endswith('.docx')):
                if HAS_DOCX:
                    try:
                        docx_file = io.BytesIO(binary_data)
                        doc = docx.Document(docx_file)
                        text_parts = [para.text for para in doc.paragraphs if para.text]
                        if text_parts:
                            text_content = "\n".join(text_parts)
                            extraction_method = "python-docx"
                    except Exception as e:
                        text_content = f"[DOCX extraction failed: {str(e)}]"
                else:
                    text_content = "[DOCX extraction requires python-docx - run: pip install python-docx]"

            # Excel files (.xlsx)
            elif 'spreadsheetml' in content_type_lower or (filename and filename.lower().endswith('.xlsx')):
                if HAS_OPENPYXL:
                    try:
                        xlsx_file = io.BytesIO(binary_data)
                        wb = openpyxl.load_workbook(xlsx_file, read_only=True, data_only=True)
                        text_parts = []
                        for sheet_name in wb.sheetnames:
                            sheet = wb[sheet_name]
                            text_parts.append(f"=== Sheet: {sheet_name} ===")
                            for row in sheet.iter_rows(values_only=True):
                                row_text = "\t".join(str(cell) if cell is not None else "" for cell in row)
                                if row_text.strip():
                                    text_parts.append(row_text)
                        wb.close()
                        if text_parts:
                            text_content = "\n".join(text_parts)
                            extraction_method = "openpyxl"
                    except Exception as e:
                        text_content = f"[XLSX extraction failed: {str(e)}]"
                else:
                    text_content = "[XLSX extraction requires openpyxl - run: pip install openpyxl]"

            # Email files (.eml)
            elif 'rfc822' in content_type_lower or (filename and filename.lower().endswith('.eml')):
                try:
                    import email
                    from email import policy
                    msg = email.message_from_bytes(binary_data, policy=policy.default)
                    text_parts = []
                    text_parts.append(f"From: {msg.get('From', '')}")
                    text_parts.append(f"To: {msg.get('To', '')}")
                    text_parts.append(f"Subject: {msg.get('Subject', '')}")
                    text_parts.append(f"Date: {msg.get('Date', '')}")
                    text_parts.append("---")
                    # Get body
                    if msg.is_multipart():
                        for part in msg.walk():
                            if part.get_content_type() == 'text/plain':
                                body = part.get_payload(decode=True)
                                if body:
                                    text_parts.append(body.decode('utf-8', errors='ignore'))
                    else:
                        body = msg.get_payload(decode=True)
                        if body:
                            text_parts.append(body.decode('utf-8', errors='ignore'))
                    text_content = "\n".join(text_parts)
                    extraction_method = "email"
                except Exception as e:
                    text_content = f"[Email extraction failed: {str(e)}]"

            # CSV files
            elif 'csv' in content_type_lower or (filename and filename.lower().endswith('.csv')):
                try:
                    text_content = binary_data.decode('utf-8')
                    extraction_method = "csv/utf-8"
                except:
                    try:
                        text_content = binary_data.decode('latin-1')
                        extraction_method = "csv/latin-1"
                    except:
                        pass

        except Exception as e:
            text_content = f"[Text extraction error: {str(e)}]"

        return text_content, extraction_method

    def _show_ai_export_options(self):
        """Show dialog for AI export configuration options"""
        dialog = tk.Toplevel(self.root)
        dialog.title("AI Export Options")
        dialog.geometry("500x620")
        dialog.transient(self.root)
        dialog.grab_set()

        # Center the dialog
        dialog.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() - 500) // 2
        y = self.root.winfo_y() + (self.root.winfo_height() - 620) // 2
        dialog.geometry(f"+{x}+{y}")

        main_frame = ttk.Frame(dialog, padding="15")
        main_frame.pack(fill="both", expand=True)

        # Title
        ttk.Label(main_frame, text="Configure AI/Copilot Export",
                 font=('TkDefaultFont', 11, 'bold')).pack(pady=(0, 15))

        # Content Options Frame
        content_frame = ttk.LabelFrame(main_frame, text="Content Options", padding="10")
        content_frame.pack(fill="x", pady=5)

        self._ai_include_email_body = tk.BooleanVar(value=True)
        ttk.Checkbutton(content_frame, text="Include email body text",
                       variable=self._ai_include_email_body).pack(anchor="w")

        self._ai_include_attachments = tk.BooleanVar(value=True)
        ttk.Checkbutton(content_frame, text="Extract text from attachments (Word, Excel, PDF, etc.)",
                       variable=self._ai_include_attachments).pack(anchor="w")

        self._ai_include_match_snippets = tk.BooleanVar(value=True)
        ttk.Checkbutton(content_frame, text="Include policy match snippets/highlights",
                       variable=self._ai_include_match_snippets).pack(anchor="w")

        self._ai_include_history = tk.BooleanVar(value=False)
        ttk.Checkbutton(content_frame, text="Include incident history",
                       variable=self._ai_include_history).pack(anchor="w")

        # Privacy/Anonymization Frame
        privacy_frame = ttk.LabelFrame(main_frame, text="Privacy Options", padding="10")
        privacy_frame.pack(fill="x", pady=5)

        self._ai_anonymize_emails = tk.BooleanVar(value=False)
        ttk.Checkbutton(privacy_frame, text="Anonymize email addresses (keep domain only: user@domain.com → @domain.com)",
                       variable=self._ai_anonymize_emails).pack(anchor="w")

        self._ai_anonymize_names = tk.BooleanVar(value=False)
        ttk.Checkbutton(privacy_frame, text="Anonymize sender/recipient names",
                       variable=self._ai_anonymize_names).pack(anchor="w")

        self._ai_anonymize_paths = tk.BooleanVar(value=False)
        ttk.Checkbutton(privacy_frame, text="Strip file paths (keep filename only)",
                       variable=self._ai_anonymize_paths).pack(anchor="w")

        # Size Control Frame
        size_frame = ttk.LabelFrame(main_frame, text="Size Control (Copilot limit: ~250MB)", padding="10")
        size_frame.pack(fill="x", pady=5)

        size_row1 = ttk.Frame(size_frame)
        size_row1.pack(fill="x", pady=2)
        ttk.Label(size_row1, text="Max file size (MB):").pack(side="left")
        self._ai_max_size_mb = tk.StringVar(value="200")
        ttk.Entry(size_row1, textvariable=self._ai_max_size_mb, width=10).pack(side="left", padx=5)
        ttk.Label(size_row1, text="(0 = unlimited, will split into multiple files)").pack(side="left")

        size_row2 = ttk.Frame(size_frame)
        size_row2.pack(fill="x", pady=2)
        ttk.Label(size_row2, text="Max incidents per file:").pack(side="left")
        self._ai_max_incidents = tk.StringVar(value="0")
        ttk.Entry(size_row2, textvariable=self._ai_max_incidents, width=10).pack(side="left", padx=5)
        ttk.Label(size_row2, text="(0 = unlimited)").pack(side="left")

        # EML Handling Frame
        eml_frame = ttk.LabelFrame(main_frame, text="Email (.eml) Processing", padding="10")
        eml_frame.pack(fill="x", pady=5)

        self._ai_parse_eml = tk.BooleanVar(value=True)
        ttk.Checkbutton(eml_frame, text="Parse .eml files and extract readable content",
                       variable=self._ai_parse_eml).pack(anchor="w")

        self._ai_include_eml_headers = tk.BooleanVar(value=True)
        ttk.Checkbutton(eml_frame, text="Include email headers (From, To, Subject, Date)",
                       variable=self._ai_include_eml_headers).pack(anchor="w")

        self._ai_include_eml_attachments = tk.BooleanVar(value=True)
        ttk.Checkbutton(eml_frame, text="Extract and include text from email attachments",
                       variable=self._ai_include_eml_attachments).pack(anchor="w")

        # Result variable
        self._ai_export_confirmed = False

        # Buttons
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill="x", pady=15)

        def on_export():
            self._ai_export_confirmed = True
            dialog.destroy()

        def on_cancel():
            self._ai_export_confirmed = False
            dialog.destroy()

        ttk.Button(btn_frame, text="Export", command=on_export, width=15).pack(side="right", padx=5)
        ttk.Button(btn_frame, text="Cancel", command=on_cancel, width=15).pack(side="right")

        # Wait for dialog to close
        self.root.wait_window(dialog)

        return self._ai_export_confirmed

    def _export_json_ai_friendly(self):
        """Export all incidents to a single JSON file optimized for AI analysis (no binary data)"""
        # Show options dialog first
        if not self._show_ai_export_options():
            return

        filepath = self.output_path_entry.get().strip()

        if not filepath:
            messagebox.showerror("Error", "Please specify an output path")
            return

        # Handle case where user provides a directory instead of filename
        if os.path.isdir(filepath) or (not filepath.lower().endswith('.json') and not os.path.splitext(filepath)[1]):
            # It's a directory or has no extension - create filename in that directory
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            os.makedirs(filepath, exist_ok=True)
            base_filename = os.path.join(filepath, f"dlp_ai_export_{timestamp}")
        else:
            # Ensure .json extension
            if filepath.lower().endswith('.json'):
                base_filename = filepath[:-5]  # Remove .json
            else:
                base_filename = filepath
            # Ensure parent directory exists
            parent_dir = os.path.dirname(base_filename)
            if parent_dir:
                os.makedirs(parent_dir, exist_ok=True)

        # Get export options
        export_options = {
            "include_email_body": self._ai_include_email_body.get(),
            "include_attachments": self._ai_include_attachments.get(),
            "include_match_snippets": self._ai_include_match_snippets.get(),
            "include_history": self._ai_include_history.get(),
            "anonymize_emails": self._ai_anonymize_emails.get(),
            "anonymize_names": self._ai_anonymize_names.get(),
            "anonymize_paths": self._ai_anonymize_paths.get(),
            "parse_eml": self._ai_parse_eml.get(),
            "include_eml_headers": self._ai_include_eml_headers.get(),
            "include_eml_attachments": self._ai_include_eml_attachments.get(),
        }

        try:
            max_size_mb = int(self._ai_max_size_mb.get() or 0)
            max_incidents = int(self._ai_max_incidents.get() or 0)
        except ValueError:
            max_size_mb = 200
            max_incidents = 0

        self.log(f"Exporting AI-friendly JSON to: {base_filename}.json")
        self.log(f"  Options: max_size={max_size_mb}MB, max_incidents={max_incidents}")

        try:
            use_cache = bool(self._fetched_incident_ids)
            if not use_cache and not self.current_incidents:
                messagebox.showinfo("Info", "No detailed incident data. Please fetch incident details first.")
                return

            # Process incidents with options
            cleaned_incidents = {}
            current_file_num = 1
            current_size_estimate = 0
            files_written = []

            if use_cache:
                total = len(self._fetched_incident_ids)
                incident_iter = enumerate(self._iter_cached_incidents())
            else:
                total = len(self.current_incidents)
                incident_iter = enumerate((inc.get("incidentId") or inc.get("id") or i, inc)
                                         for i, inc in enumerate(self.current_incidents))

            for i, (incident_id, details) in incident_iter:
                cleaned = self._clean_incident_for_ai(details, export_options)
                cleaned_incidents[str(incident_id)] = cleaned

                # Estimate size (rough: 1 char ≈ 1 byte)
                incident_size = len(json.dumps(cleaned, default=str))
                current_size_estimate += incident_size

                if (i + 1) % 100 == 0:
                    self.log(f"  Processed {i + 1}/{total} incidents (~{current_size_estimate // (1024*1024)}MB)...")
                    self.root.update_idletasks()

                # Check if we need to split
                should_split = False
                if max_size_mb > 0 and current_size_estimate > max_size_mb * 1024 * 1024:
                    should_split = True
                if max_incidents > 0 and len(cleaned_incidents) >= max_incidents:
                    should_split = True

                if should_split and len(cleaned_incidents) > 1:
                    # Write current batch
                    filename = f"{base_filename}_part{current_file_num}.json" if max_size_mb > 0 or max_incidents > 0 else f"{base_filename}.json"
                    self._write_ai_export_file(filename, cleaned_incidents, export_options, current_file_num, total)
                    files_written.append(filename)

                    # Reset for next batch
                    cleaned_incidents = {}
                    current_size_estimate = 0
                    current_file_num += 1

            # Write remaining incidents
            if cleaned_incidents:
                if files_written:
                    filename = f"{base_filename}_part{current_file_num}.json"
                else:
                    filename = f"{base_filename}.json"
                self._write_ai_export_file(filename, cleaned_incidents, export_options, current_file_num if files_written else None, total)
                files_written.append(filename)

            # Summary
            total_size = sum(os.path.getsize(f) for f in files_written)
            self.log(f"Export complete: {len(files_written)} file(s), {total_size:,} bytes total", "SUCCESS")

            if len(files_written) == 1:
                messagebox.showinfo("Export Complete",
                                  f"Exported to:\n{files_written[0]}\n\n"
                                  f"File size: {total_size:,} bytes\n\n"
                                  f"Ready for AI/Copilot analysis.")
            else:
                messagebox.showinfo("Export Complete",
                                  f"Exported to {len(files_written)} files:\n"
                                  f"{os.path.dirname(files_written[0])}\n\n"
                                  f"Total size: {total_size:,} bytes\n\n"
                                  f"Files split to stay under {max_size_mb}MB limit.")

        except Exception as e:
            self.log(f"Export failed: {str(e)}", "ERROR")
            import traceback
            self.log(f"Traceback: {traceback.format_exc()}", "ERROR")
            messagebox.showerror("Export Failed", str(e))

    def _write_ai_export_file(self, filename, incidents, options, part_num, total_incidents):
        """Write a single AI export file"""
        export_data = {
            "exportDate": datetime.now().isoformat(),
            "exportType": "AI-Friendly (Copilot Compatible)",
            "exportOptions": options,
            "environment": self.env_var.get(),
            "server": self.server_entry.get(),
            "partNumber": part_num,
            "incidentsInFile": len(incidents),
            "totalIncidentsInExport": total_incidents,
            "incidents": incidents
        }

        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2, default=str)

        file_size = os.path.getsize(filename)
        self.log(f"  Written: {os.path.basename(filename)} ({file_size:,} bytes, {len(incidents)} incidents)")

    def _clean_incident_for_ai(self, details, options=None):
        """Remove binary data from incident details, extract text content where possible"""
        import base64
        import re

        # Default options if not provided
        if options is None:
            options = {
                "include_email_body": True,
                "include_attachments": True,
                "include_match_snippets": True,
                "include_history": False,
                "anonymize_emails": False,
                "anonymize_names": False,
                "anonymize_paths": False,
                "parse_eml": True,
                "include_eml_headers": True,
                "include_eml_attachments": True,
            }

        cleaned = {}

        # Build component names map for better identification
        component_names = {}
        if "staticAttributes" in details and isinstance(details["staticAttributes"], dict):
            info_map = details["staticAttributes"].get("infoMap", {})
            if isinstance(info_map, dict):
                attachment_info = info_map.get("attachmentInfo", [])
                if isinstance(attachment_info, list):
                    for att in attachment_info:
                        if isinstance(att, dict):
                            comp_id = att.get("componentId") or att.get("id")
                            att_name = att.get("attachmentName") or att.get("name")
                            if comp_id and att_name:
                                component_names[str(comp_id)] = att_name

        def anonymize_email(text):
            """Replace email addresses with domain only"""
            if not options.get("anonymize_emails"):
                return text
            if not isinstance(text, str):
                return text
            # Replace user@domain.com with @domain.com
            return re.sub(r'[a-zA-Z0-9._%+-]+@([a-zA-Z0-9.-]+\.[a-zA-Z]{2,})', r'@\1', text)

        def anonymize_text(text):
            """Apply all anonymization options to text"""
            if not isinstance(text, str):
                return text
            result = text
            if options.get("anonymize_emails"):
                result = anonymize_email(result)
            if options.get("anonymize_paths"):
                # Keep only filename from paths
                result = re.sub(r'[A-Za-z]:\\[^\s"\'<>|]*\\([^\\"\s\'<>|]+)', r'\1', result)
                result = re.sub(r'/[^\s"\'<>|]*/([^/"\s\'<>|]+)', r'\1', result)
            return result

        def process_dict_recursive(obj):
            """Recursively process dict/list to anonymize strings"""
            if isinstance(obj, dict):
                return {k: process_dict_recursive(v) for k, v in obj.items()}
            elif isinstance(obj, list):
                return [process_dict_recursive(item) for item in obj]
            elif isinstance(obj, str):
                return anonymize_text(obj)
            return obj

        for key, value in details.items():
            if key == "originalMessage":
                if not options.get("include_email_body"):
                    cleaned[key] = {"skipped": "Email body excluded by export options"}
                    continue

                # Extract text from email body if possible
                if isinstance(value, dict) and "dataBase64" in value:
                    content_type = value.get("contentType", "")
                    try:
                        binary_data = base64.b64decode(value["dataBase64"])

                        # Check if it's an EML file and parse it specially
                        detected_type = detect_file_type_by_magic(binary_data)
                        if detected_type == '.eml' and options.get("parse_eml"):
                            parsed_email = self._parse_eml_for_ai(binary_data, options)
                            cleaned[key] = {
                                "contentType": content_type,
                                "detectedType": ".eml",
                                "size": value.get("size"),
                                "parsedEmail": process_dict_recursive(parsed_email)
                            }
                        else:
                            extracted_text, method = self._extract_text_from_binary(binary_data, content_type, "email.eml")
                            cleaned[key] = {
                                "contentType": content_type,
                                "size": value.get("size"),
                                "extractedText": anonymize_text(extracted_text),
                                "extractionMethod": method
                            }
                    except Exception as e:
                        cleaned[key] = {
                            "contentType": content_type,
                            "size": value.get("size"),
                            "extractionError": str(e)
                        }
                else:
                    cleaned[key] = process_dict_recursive(value)

            elif key == "componentData":
                if not options.get("include_attachments"):
                    cleaned[key] = {"skipped": "Attachments excluded by export options"}
                    continue

                # Extract text from attachments where possible
                if isinstance(value, dict):
                    cleaned[key] = {}
                    for comp_id, comp_data in value.items():
                        if isinstance(comp_data, dict) and "dataBase64" in comp_data:
                            content_type = comp_data.get("contentType", "")
                            filename = component_names.get(str(comp_id), f"attachment_{comp_id}")
                            if options.get("anonymize_paths"):
                                filename = os.path.basename(filename)

                            try:
                                binary_data = base64.b64decode(comp_data["dataBase64"])

                                # Check if it's an EML file
                                detected_type = detect_file_type_by_magic(binary_data)
                                if detected_type == '.eml' and options.get("parse_eml"):
                                    parsed_email = self._parse_eml_for_ai(binary_data, options)
                                    cleaned[key][comp_id] = {
                                        "filename": filename,
                                        "contentType": content_type,
                                        "detectedType": ".eml",
                                        "size": comp_data.get("size"),
                                        "parsedEmail": process_dict_recursive(parsed_email)
                                    }
                                else:
                                    extracted_text, method = self._extract_text_from_binary(binary_data, content_type, filename)
                                    cleaned[key][comp_id] = {
                                        "filename": filename,
                                        "contentType": content_type,
                                        "detectedType": detected_type,
                                        "size": comp_data.get("size"),
                                        "extractedText": anonymize_text(extracted_text),
                                        "extractionMethod": method
                                    }
                            except Exception as e:
                                cleaned[key][comp_id] = {
                                    "filename": filename,
                                    "contentType": content_type,
                                    "size": comp_data.get("size"),
                                    "extractionError": str(e)
                                }
                        else:
                            cleaned[key][comp_id] = process_dict_recursive(comp_data)
                else:
                    cleaned[key] = process_dict_recursive(value)

            elif key == "history":
                if options.get("include_history"):
                    cleaned[key] = process_dict_recursive(value)
                # else skip it

            elif key in ("componentMatches", "matches"):
                if options.get("include_match_snippets"):
                    cleaned[key] = process_dict_recursive(value)
                else:
                    # Keep structure but remove match text snippets
                    if isinstance(value, list):
                        cleaned[key] = []
                        for item in value:
                            if isinstance(item, dict):
                                filtered = {k: v for k, v in item.items()
                                          if k not in ("matchedContent", "matchText", "highlightedText", "snippet")}
                                cleaned[key].append(process_dict_recursive(filtered))
                            else:
                                cleaned[key].append(item)
                    else:
                        cleaned[key] = process_dict_recursive(value)

            elif key in ("components", "policyMatches", "staticAttributes",
                        "editableAttributes", "correlations", "incidentMetadata"):
                # Keep all text-based detail sections, apply anonymization
                cleaned[key] = process_dict_recursive(value)

            else:
                # Keep other fields
                cleaned[key] = process_dict_recursive(value)

        return cleaned

    def _parse_eml_for_ai(self, eml_data, options):
        """Parse EML file and extract content in AI-friendly format"""
        import email
        from email import policy
        import base64

        result = {
            "headers": {},
            "body": {},
            "attachments": []
        }

        try:
            # Parse the email
            if isinstance(eml_data, bytes):
                msg = email.message_from_bytes(eml_data, policy=policy.default)
            else:
                msg = email.message_from_string(eml_data, policy=policy.default)

            # Extract headers if requested
            if options.get("include_eml_headers", True):
                for header in ["From", "To", "Cc", "Bcc", "Subject", "Date", "Reply-To", "Message-ID"]:
                    if msg[header]:
                        result["headers"][header] = str(msg[header])

            # Extract body
            if msg.is_multipart():
                for part in msg.walk():
                    content_type = part.get_content_type()
                    content_disposition = str(part.get("Content-Disposition", ""))

                    # Skip attachments in first pass
                    if "attachment" in content_disposition:
                        continue

                    if content_type == "text/plain":
                        try:
                            body_text = part.get_content()
                            if isinstance(body_text, bytes):
                                body_text = body_text.decode('utf-8', errors='replace')
                            result["body"]["plain"] = body_text
                        except:
                            pass
                    elif content_type == "text/html":
                        try:
                            html_content = part.get_content()
                            if isinstance(html_content, bytes):
                                html_content = html_content.decode('utf-8', errors='replace')
                            # Strip HTML tags for AI readability
                            import re
                            text = re.sub(r'<style[^>]*>.*?</style>', '', html_content, flags=re.DOTALL | re.IGNORECASE)
                            text = re.sub(r'<script[^>]*>.*?</script>', '', text, flags=re.DOTALL | re.IGNORECASE)
                            text = re.sub(r'<[^>]+>', ' ', text)
                            text = re.sub(r'\s+', ' ', text).strip()
                            # Decode HTML entities
                            import html
                            text = html.unescape(text)
                            result["body"]["html_as_text"] = text
                        except:
                            pass

                # Extract attachments if requested
                if options.get("include_eml_attachments", True):
                    for part in msg.walk():
                        content_disposition = str(part.get("Content-Disposition", ""))
                        if "attachment" in content_disposition:
                            filename = part.get_filename() or "unnamed_attachment"
                            content_type = part.get_content_type()

                            try:
                                payload = part.get_payload(decode=True)
                                if payload:
                                    # Try to extract text from the attachment
                                    extracted_text, method = self._extract_text_from_binary(
                                        payload, content_type, filename
                                    )
                                    result["attachments"].append({
                                        "filename": filename,
                                        "contentType": content_type,
                                        "size": len(payload),
                                        "extractedText": extracted_text,
                                        "extractionMethod": method
                                    })
                            except Exception as e:
                                result["attachments"].append({
                                    "filename": filename,
                                    "contentType": content_type,
                                    "extractionError": str(e)
                                })
            else:
                # Not multipart - simple email
                content_type = msg.get_content_type()
                try:
                    body = msg.get_content()
                    if isinstance(body, bytes):
                        body = body.decode('utf-8', errors='replace')
                    result["body"]["content"] = body
                except:
                    result["body"]["content"] = str(msg.get_payload())

        except Exception as e:
            result["parseError"] = str(e)
            # Fall back to raw text
            try:
                if isinstance(eml_data, bytes):
                    result["rawText"] = eml_data.decode('utf-8', errors='replace')
                else:
                    result["rawText"] = str(eml_data)
            except:
                pass

        return result

    def _export_csv(self):
        """Export full incident details to CSV"""
        filename = self.output_path_entry.get().strip()

        if not filename:
            messagebox.showerror("Error", "Please specify an output path")
            return

        # Ensure .csv extension
        if not filename.lower().endswith('.csv'):
            filename += '.csv'

        self.log(f"Exporting CSV to: {filename}")

        try:
            use_cache = bool(self._fetched_incident_ids)

            # Create iterator based on data source
            if use_cache:
                incident_iterator = self._iter_cached_incidents()
                total = len(self._fetched_incident_ids)
            else:
                incident_iterator = ((inc.get("incidentId") or inc.get("id") or i, inc)
                                    for i, inc in enumerate(self.current_incidents))
                total = len(self.current_incidents)

            rows = []
            processed = 0
            for incident_id, details in incident_iterator:
                processed += 1
                if processed % 100 == 0:
                    self.log(f"  Processing {processed}/{total} incidents for CSV...")
                    self.root.update_idletasks()
                row = {"incidentId": incident_id}

                # Extract from staticAttributes - including nested infoMap
                if "staticAttributes" in details and details["staticAttributes"]:
                    static = details["staticAttributes"]
                    if isinstance(static, dict):
                        # Get top-level static attributes
                        for key, value in static.items():
                            if key == "infoMap" and isinstance(value, dict):
                                # Flatten infoMap
                                for info_key, info_value in value.items():
                                    if isinstance(info_value, list):
                                        # Handle arrays like recipientInfo, attachmentInfo
                                        if info_key == "recipientInfo" and info_value:
                                            recipients = [r.get("recipientIdentifier", "") for r in info_value if isinstance(r, dict)]
                                            row["recipients"] = "; ".join(recipients)
                                        elif info_key == "attachmentInfo" and info_value:
                                            attachments = [a.get("attachmentName", "") for a in info_value if isinstance(a, dict)]
                                            row["attachments"] = "; ".join(attachments)
                                            sizes = [a.get("attachmentSize", "") for a in info_value if isinstance(a, dict)]
                                            row["attachmentSizes"] = "; ".join(sizes)
                                        else:
                                            row[info_key] = json.dumps(info_value)
                                    elif isinstance(info_value, dict):
                                        row[info_key] = json.dumps(info_value)
                                    else:
                                        row[info_key] = info_value
                            elif not isinstance(value, (dict, list)):
                                row[key] = value

                # Extract editable attributes
                if "editableAttributes" in details and details["editableAttributes"]:
                    editable = details["editableAttributes"]
                    if isinstance(editable, dict):
                        for key, value in editable.items():
                            if not isinstance(value, (dict, list)):
                                row[f"editable_{key}"] = value

                # Extract component matches (violations)
                if "componentMatches" in details and details["componentMatches"]:
                    comp_matches = details["componentMatches"]
                    if isinstance(comp_matches, list):
                        # Collect all violations
                        all_rules = []
                        all_violations = []
                        for comp in comp_matches:
                            if isinstance(comp, dict):
                                comp_name = comp.get("messageComponentName", "")
                                violations = comp.get("violations", [])
                                for violation in violations:
                                    if isinstance(violation, dict):
                                        segments = violation.get("violationSegments", [])
                                        for seg in segments:
                                            if isinstance(seg, dict):
                                                rule_name = seg.get("ruleName", "")
                                                text = seg.get("text", "")
                                                if rule_name:
                                                    all_rules.append(rule_name)
                                                if text:
                                                    all_violations.append(f"{rule_name}: {text}")
                        row["ruleNames"] = "; ".join(set(all_rules))
                        row["violationDetails"] = "; ".join(all_violations)

                # Extract policy matches
                if "policyMatches" in details and details["policyMatches"]:
                    policy_matches = details["policyMatches"]
                    if isinstance(policy_matches, list):
                        policies = [p.get("policyName", "") for p in policy_matches if isinstance(p, dict)]
                        row["violatedPolicies"] = "; ".join(policies)

                # Add counts
                if "components" in details and isinstance(details["components"], list):
                    row["componentCount"] = len(details["components"])
                if "matches" in details and isinstance(details["matches"], list):
                    row["matchCount_fromAPI"] = len(details["matches"])
                if "correlations" in details and isinstance(details["correlations"], list):
                    row["correlationCount"] = len(details["correlations"])

                # Extract history summary
                if "history" in details and details["history"]:
                    history = details["history"]
                    if isinstance(history, list) and history:
                        # Get last history entry
                        last_entry = history[0] if history else {}
                        if isinstance(last_entry, dict):
                            row["lastHistoryAction"] = last_entry.get("action", "")
                            row["lastHistoryDate"] = last_entry.get("date", "")
                            row["lastHistoryUser"] = last_entry.get("userName", "")

                rows.append(row)

            if rows:
                # Define preferred column order
                priority_columns = [
                    "incidentId", "policyName", "policyId", "messageTypeName", "messageType",
                    "severityId", "matchCount", "detectionDate", "messageDate", "creationDate",
                    "domainUserName", "networkSenderIdentifier", "senderIPAddress",
                    "recipients", "endpointMachineName", "endpointMachineIpAddress",
                    "endpointApplicationName", "attachments", "attachmentSizes",
                    "ruleNames", "violationDetails", "violatedPolicies",
                    "endpointJustificationLabel", "endpointJustificationUserText",
                    "detectionServerName", "incidentStatusName"
                ]

                # Get all unique columns
                all_columns = set()
                for row in rows:
                    all_columns.update(row.keys())

                # Order columns: priority first, then rest alphabetically
                ordered_columns = []
                for col in priority_columns:
                    if col in all_columns:
                        ordered_columns.append(col)
                        all_columns.discard(col)
                ordered_columns.extend(sorted(all_columns))

                with open(filename, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.DictWriter(f, fieldnames=ordered_columns)
                    writer.writeheader()
                    writer.writerows(rows)

                self.log(f"Exported CSV with {len(ordered_columns)} columns to {filename}", "SUCCESS")
                messagebox.showinfo("Export Complete", f"Exported {len(rows)} incidents with {len(ordered_columns)} columns to:\n{filename}")
            else:
                messagebox.showinfo("Info", "No data to export. Try fetching incident details first.")

        except Exception as e:
            self.log(f"CSV export failed: {str(e)}", "ERROR")
            import traceback
            self.log(f"Traceback: {traceback.format_exc()}", "ERROR")
            messagebox.showerror("Export Failed", str(e))

    def _export_individual_json(self):
        """Export each incident to individual JSON files"""
        use_cache = bool(self._fetched_incident_ids)
        if not use_cache and not self.current_incidents:
            messagebox.showinfo("Info", "No detailed incident data. Please fetch incident details first.")
            return

        folder = self.output_path_entry.get().strip()

        if not folder:
            messagebox.showerror("Error", "Please specify an output folder path")
            return

        self.log(f"Exporting individual JSON files to: {folder}")

        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            export_folder = os.path.join(folder, f"dlp_incidents_{timestamp}")
            os.makedirs(export_folder, exist_ok=True)

            # Create iterator based on data source
            if use_cache:
                incident_iterator = self._iter_cached_incidents()
                total = len(self._fetched_incident_ids)
            else:
                incident_iterator = ((inc.get("incidentId") or inc.get("id") or i, inc)
                                    for i, inc in enumerate(self.current_incidents))
                total = len(self.current_incidents)

            exported = 0
            for incident_id, details in incident_iterator:
                filename = os.path.join(export_folder, f"incident_{incident_id}.json")
                with open(filename, 'w', encoding='utf-8') as f:
                    json.dump(details, f, indent=2, default=str)
                exported += 1
                if exported % 100 == 0:
                    self.log(f"  Exported {exported}/{total} individual JSON files...")
                    self.root.update_idletasks()

            self.log(f"Exported {exported} files to {export_folder}", "SUCCESS")
            messagebox.showinfo("Export Complete",
                              f"Exported {exported} incident files to:\n{export_folder}")

        except Exception as e:
            self.log(f"Export failed: {str(e)}", "ERROR")
            messagebox.showerror("Export Failed", str(e))


def main():
    """Main entry point"""
    root = tk.Tk()
    app = DLPIncidentDownloader(root)
    root.mainloop()


if __name__ == "__main__":
    main()
